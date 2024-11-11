import os
import re
import json
import sqlite3
import asyncio
import logging
from flask import Flask, render_template, request, send_file, redirect, session, current_app, make_response, after_this_request, Response, stream_with_context, flash, abort, redirect, send_from_directory
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from werkzeug.utils import secure_filename
from rapidfuzz import process, fuzz
from collections import defaultdict, Counter
from typing import List, Set, Dict, Tuple, Iterator, Set, Callable
import tempfile
import openpyxl
from openpyxl.styles import NamedStyle
from tqdm import tqdm
import http.client as http_client
import time
import shutil
import gc
from concurrent.futures import ThreadPoolExecutor, as_completed
import numpy as np
from werkzeug.utils import secure_filename
import plotly.express as px
from sentence_transformers import SentenceTransformer
from sklearn.cluster import KMeans
from sklearn.manifold import TSNE
from sklearn.decomposition import PCA
from sklearn.metrics import silhouette_score
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import threading
import oci
from oci.generative_ai_inference import GenerativeAiInferenceClient
from oci.generative_ai_inference.models import ChatDetails, CohereChatRequest, OnDemandServingMode
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
from functools import wraps
from pathlib import Path
import io
from io import BytesIO
import plotly.io as pio
from datetime import datetime

from config import OCI_CONFIG_PATH, OCI_CONFIG_PROFILE, GENERATIVE_AI_MODEL_ID, SERVICE_ENDPOINT, NAMESPACE, BUCKET_NAME, COMPARTMENT_NAME

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Download required NLTK data
nltk.download('stopwords')
nltk.download('wordnet')
stop_words = set(stopwords.words('english'))
lemmatizer = WordNetLemmatizer()

# Set up pymedtermino before importing SNOMEDCT
import pymedtermino

# Set the DATA_DIR for pymedtermino
pymedtermino.DATA_DIR = "/home/opc"

# Corrected: Remove the .sqlite3 extension and store the connection
db = pymedtermino.connect_sqlite3("/home/opc/snomedct")

# Now import SNOMEDCT after setting DATA_DIR and connecting
from pymedtermino.snomedct import SNOMEDCT

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Download required NLTK data
nltk.download('stopwords')
nltk.download('wordnet')
stop_words = set(stopwords.words('english'))
lemmatizer = WordNetLemmatizer()

config_path = '/home/opc/.oci/config'
CONFIG_PROFILE = "DEFAULT"
endpoint = "https://inference.generativeai.us-chicago-1.oci.oraclecloud.com"

try:
    config = oci.config.from_file(config_path, CONFIG_PROFILE)
    logging.info("OCI configuration loaded successfully.")
except Exception as e:
    logging.error(f"Error loading OCI configuration: {str(e)}")
    exit(1)

compartment_id = config['tenancy']

# Initialize the GenerativeAiInferenceClient with the ConnectionPoolConfig
generative_ai_inference_client = GenerativeAiInferenceClient(
    config=config,
    service_endpoint=endpoint,
    retry_strategy=oci.retry.NoneRetryStrategy(),
    timeout=(10, 240),
)

# Limit the number of concurrent threads
max_workers = 15
batch_size = 15

semaphore = threading.Semaphore(10)

log_filename = f'logs/embedding_process_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()  # This will show logs in git bash/console
    ]
)

app = Flask(__name__)
app.secret_key = 'alsdkfja2342342adflkadjf' 

UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed_files'
ALLOWED_EXTENSIONS = {'xlsx'}
RESULTS_FOLDER = 'results'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['RESULTS_FOLDER'] = RESULTS_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

# Ensure the upload and processed directories exist
os.makedirs('uploads', exist_ok=True)
os.makedirs('processed_files', exist_ok=True)

# Helper function to check file extension
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Home route for the web hub
@app.route('/')
def home():
    return render_template('home.html')

CUSTOM_DICT_FILE = "custom_dictionary.json"

# Load SNOMED words once at startup
print("Loading SNOMED CT terms...")
snomed_words = set()

def get_snomed_words():
    # Use the existing database connection
    cursor = db.cursor()
    
    # Query the active descriptions from the database
    cursor.execute("SELECT term FROM Descriptions WHERE active=1")
    snomed_terms = cursor.fetchall()
    cursor.close()
    
    # Extract words and convert them to a set
    words = set()
    for term in snomed_terms:
        if term[0]:
            words.update(re.findall(r'\b\w+\b', term[0].lower()))
    return words

snomed_words = get_snomed_words()
print(f"Loaded {len(snomed_words)} unique words from SNOMED CT")

# Utility functions (from your script, adjusted for web context)
def load_custom_dictionary():
    try:
        with open(CUSTOM_DICT_FILE, 'r') as f:
            return set(json.load(f))
    except FileNotFoundError:
        return set()

def save_custom_dictionary(custom_dict):
    with open(CUSTOM_DICT_FILE, 'w') as f:
        json.dump(list(custom_dict), f)

def match_case(word, corrected_word):
    if word.isupper():
        return corrected_word.upper()
    elif word.islower():
        return corrected_word.lower()
    elif word.istitle():
        return corrected_word.capitalize()
    else:
        return corrected_word

def process_alias_word(word, custom_words):
    # Replace '*' and '-' with '_'
    word = word.replace('*', '_')
    word = word.replace('-', '_')
    
    # Check if the word is in the custom dictionary (case-insensitive)
    if word in custom_words or word.lower() in custom_words:
        return word.upper()
    # Handle camel case words for alias, and convert to upper case
    elif re.search(r'[a-z][A-Z]', word):
        return re.sub(r'([a-z])([A-Z])', r'\1_\2', word).upper()
    else:
        return word.upper()

def convert_name(name, append_text, use_spellcheck=True, custom_dict=set(), suppress_final_output=False):
    # Initialize words
    custom_words = set(['IgM', 'IgG', 'IgD', 'IgE', 'IgA', 'CSF', 'rRNA', 'DNA', 'mRNA'])
    custom_words.update(custom_dict)
    
    # Create a combined set of words for spell-checking
    combined_words = snomed_words.union(word.lower() for word in custom_words)

    corrected_words = []
    # Split the name into words and punctuation
    words = re.findall(r'\w+|\s+|[^\w\s]+', name)

    for original_word in words:
        if not original_word.strip() or not use_spellcheck or not original_word.isalpha():
            corrected_words.append(original_word)
        elif original_word.lower() in combined_words or original_word in custom_words:
            corrected_words.append(original_word)
        else:
            # Spellcheck logic
            matches = process.extract(original_word.lower(), combined_words, limit=1, scorer=fuzz.ratio)
            if matches and matches[0][1] >= 80:
                suggested_word = matches[0][0]
                if original_word.lower() != suggested_word and not re.search(r'\d', original_word):
                    # Here, instead of input(), we need to handle this interaction in the web interface
                    # We'll store the necessary information in session and prompt the user
                    session['spellcheck_needed'] = True
                    session['original_word'] = original_word
                    session['suggested_word'] = suggested_word
                    session['pending_name'] = name
                    session['corrected_words'] = corrected_words
                    session['words_remaining'] = words[words.index(original_word)+1:]
                    session['append_text'] = append_text
                    session['use_spellcheck'] = use_spellcheck
                    session['custom_dict'] = list(custom_dict)
                    return None, None  # Indicate that spellcheck interaction is needed
                else:
                    corrected_words.append(original_word)
            else:
                corrected_words.append(original_word)

    # Join words for the name, preserving original spacing and punctuation
    processed_name = ''.join(corrected_words)

    # Now build alias_name from processed_name
    alias_name = processed_name

    # Remove text within parentheses for alias processing
    alias_name = re.sub(r'\([^)]*\)', '', alias_name)
    alias_name = re.sub(r"['+/,<>:\.\-]", ' ', alias_name)
    alias_name = alias_name.replace('%', ' PCT')

    # Split alias_name into words
    alias_words = re.findall(r'\b[\w*-]+\b', alias_name)

    # Process alias words
    alias_corrected_words = []
    for alias_word in alias_words:
        alias_corrected_words.append(process_alias_word(alias_word, custom_words))

    # Join alias words to form the alias
    processed_alias = '_'.join(filter(bool, alias_corrected_words))

    # Append text to the end of the alias
    if append_text.upper() not in processed_alias:
        processed_alias = f"{processed_alias.rstrip('_')}_{append_text.upper()}"

    return processed_name, processed_alias

#stop undoing here
@app.route('/alias-converter', methods=['GET', 'POST'])
def alias_converter():
    if 'history' not in session:
        session['history'] = []
        session['state'] = 'start'

    history = session['history']
    state = session['state']

    if request.method == 'POST':
        user_input = request.form['user_input'].strip()
        if user_input.lower() == 'reset':
            session.clear()
            history = []
            history.append("Do you want to use spell check? (y/n):")
            session['state'] = 'start'
            session['history'] = history
            return render_template('terminal.html', history=history)

        history.append(f">>> {user_input}")

        # Handle the different states
        if state == 'start':
            session['use_spellcheck'] = user_input.lower() == 'y'
            session['state'] = 'append_text'
            history.append("Enter the text you want to append (e.g., OBSTYPE):")
        elif state == 'append_text':
            session['append_text'] = user_input
            session['state'] = 'add_mode'
            history.append("Do you want to perform a single add or multi add? (s/m) (or type 'exit' to quit):")
        elif state == 'add_mode':
            if user_input.lower() == 's':
                session['state'] = 'single_add'
                history.append("Enter the name you want to convert (or type 'back' to return to the main menu):")
            elif user_input.lower() == 'm':
                session['state'] = 'multi_add'
                history.append("Enter the names you want to convert, one per line (type 'done' on a new line to finish or 'back' to return to the main menu):")
            elif user_input.lower() == 'exit':
                # Clear the session and reset to start
                session.clear()
                history = []
                history.append("Do you want to use spell check? (y/n):")
                session['state'] = 'start'
                session['history'] = history
            else:
                history.append("Invalid selection. Please choose 's' for single add, 'm' for multi add, or 'exit' to quit.")
        elif state == 'single_add':
            if user_input.lower() == 'back':
                session['state'] = 'add_mode'
                history.append("Do you want to perform a single add or multi add? (s/m) (or type 'exit' to quit):")
            else:
                # Process the single name
                custom_dict = load_custom_dictionary()
                result = convert_name(
                    user_input,
                    session['append_text'],
                    session['use_spellcheck'],
                    custom_dict=custom_dict
                )
                if result == (None, None):
                    # Spellcheck interaction needed
                    history.append(f"Potential misspelling detected: '{session['original_word']}'")
                    history.append(f"Would you like to update it to '{session['suggested_word']}'? (y/n):")
                    session['state'] = 'spellcheck_prompt_single'
                else:
                    processed_name, processed_alias = result
                    history.append("Converted Names:")
                    history.append(processed_name)
                    history.append("Converted Aliases:")
                    history.append(processed_alias)
                    # Remain in single_add state
                    history.append("Enter the name you want to convert (or type 'back' to return to the main menu):")
                    session['state'] = 'single_add'
        elif state == 'multi_add':
            if user_input.lower() == 'back':
                session['state'] = 'add_mode'
                history.append("Do you want to perform a single add or multi add? (s/m) (or type 'exit' to quit):")
            elif user_input.lower() == 'done':
                # Move to processing state
                session['state'] = 'processing_multi_add'
                return alias_converter()
            else:
                # Split the input into lines
                names = user_input.strip().splitlines()
                names = [name.strip() for name in names if name.strip()]
                if not names:
                    history.append("No names entered. Please enter at least one name or type 'back' to return.")
                else:
                    # Initialize variables
                    session['remaining_names'] = names
                    session['converted_names'] = []
                    session['converted_aliases'] = []
                    session['state'] = 'processing_multi_add'
                    # Start processing
                    return alias_converter()
        elif state == 'processing_multi_add':
            # Process names in a loop
            remaining_names = session.get('remaining_names', [])
            converted_names = session.get('converted_names', [])
            converted_aliases = session.get('converted_aliases', [])
            custom_dict = load_custom_dictionary()
            spellcheck_needed = False

            while remaining_names and not spellcheck_needed:
                name = remaining_names.pop(0)
                session['remaining_names'] = remaining_names

                # Initialize variables for the new name
                session['corrected_words'] = []
                session['words_remaining'] = []
                session['original_word'] = ''
                session['suggested_word'] = ''
                session['current_name'] = name

                result = convert_name(
                    name,
                    session['append_text'],
                    session['use_spellcheck'],
                    custom_dict=custom_dict,
                    suppress_final_output=True
                )

                if result == (None, None):
                    # Spellcheck interaction needed
                    history.append(f"Potential misspelling detected: '{session['original_word']}' in '{name}'")
                    history.append(f"Would you like to update it to '{session['suggested_word']}'? (y/n):")
                    session['state'] = 'spellcheck_prompt_multi'
                    spellcheck_needed = True
                    break  # Exit the loop to wait for user input
                else:
                    processed_name, processed_alias = result
                    converted_names.append(processed_name)
                    converted_aliases.append(processed_alias)
                    session['converted_names'] = converted_names
                    session['converted_aliases'] = converted_aliases

            if not spellcheck_needed and not remaining_names:
                # All names processed, display results
                history.append("Converted Names:")
                history.append("\n".join(converted_names))
                history.append("Converted Aliases:")
                history.append("\n".join(converted_aliases))
                # Loop back to add_mode
                history.append("Do you want to perform a single add or multi add? (s/m) (or type 'exit' to quit):")
                session['state'] = 'add_mode'

            return render_template('terminal.html', history=history)
        elif state == 'spellcheck_prompt_single':
            # Handle user response to spellcheck for single add
            if user_input.lower() == 'y':
                # User accepted the suggestion
                corrected_word = match_case(session['original_word'], session['suggested_word'])
                session['corrected_words'].append(corrected_word)
            else:
                # User declined the suggestion, use the original word
                session['corrected_words'].append(session['original_word'])

            # Continue processing the remaining words
            name = session['pending_name']
            words_remaining = session.get('words_remaining', [])
            corrected_words = session.get('corrected_words', [])
            custom_dict = set(session.get('custom_dict', []))

            while words_remaining:
                original_word = words_remaining.pop(0)
                session['words_remaining'] = words_remaining
                if not original_word.strip() or not session['use_spellcheck'] or not original_word.isalpha():
                    corrected_words.append(original_word)
                elif original_word.lower() in snomed_words or original_word in custom_dict:
                    corrected_words.append(original_word)
                else:
                    # Check for potential misspelling
                    combined_words = snomed_words.union(word.lower() for word in custom_dict)
                    matches = process.extract(original_word.lower(), combined_words, limit=1, scorer=fuzz.ratio)
                    if matches and matches[0][1] >= 80:
                        suggested_word = matches[0][0]
                        if original_word.lower() != suggested_word and not re.search(r'\d', original_word):
                            # Need to prompt user for this word
                            session['original_word'] = original_word
                            session['suggested_word'] = suggested_word
                            # Update the state to prompt for this word
                            history.append(f"Potential misspelling detected: '{session['original_word']}'")
                            history.append(f"Would you like to update it to '{session['suggested_word']}'? (y/n):")
                            return render_template('terminal.html', history=history)
                        else:
                            corrected_words.append(original_word)
                    else:
                        corrected_words.append(original_word)

            # Generate the corrected name and alias after processing all words
            processed_name = ''.join(corrected_words)
            alias_name = re.sub(r'\([^)]*\)', '', processed_name)
            alias_name = re.sub(r"['+/,<>:\.\-]", ' ', alias_name)
            alias_name = alias_name.replace('%', ' PCT')
            alias_words = re.findall(r'\b[\w*-]+\b', alias_name)
            alias_corrected_words = [process_alias_word(word, custom_dict) for word in alias_words]
            processed_alias = '_'.join(filter(bool, alias_corrected_words))
            if session['append_text'].upper() not in processed_alias:
                processed_alias = f"{processed_alias.rstrip('_')}_{session['append_text'].upper()}"

            # Display the final corrected name and alias
            history.append("Converted Names:")
            history.append(processed_name)
            history.append("Converted Aliases:")
            history.append(processed_alias)

            # Remain in single_add state
            history.append("Enter the name you want to convert (or type 'back' to return to the main menu):")
            session['state'] = 'single_add'

        elif state == 'spellcheck_prompt_multi':
            if user_input.lower() == 'y':
                corrected_word = match_case(session['original_word'], session['suggested_word'])
            else:
                corrected_word = session['original_word']

            # Append corrected word to corrected_words
            corrected_words = session.get('corrected_words', [])
            corrected_words.append(corrected_word)
            session['corrected_words'] = corrected_words

            words_remaining = session.get('words_remaining', [])
            custom_dict = set(session.get('custom_dict', []))

            while words_remaining:
                original_word = words_remaining.pop(0)
                session['words_remaining'] = words_remaining
                if not original_word.strip() or not session['use_spellcheck'] or not original_word.isalpha():
                    corrected_words.append(original_word)
                elif original_word.lower() in snomed_words or original_word in custom_dict:
                    corrected_words.append(original_word)
                else:
                    # Spellcheck logic
                    combined_words = snomed_words.union(word.lower() for word in custom_dict)
                    matches = process.extract(original_word.lower(), combined_words, limit=1, scorer=fuzz.ratio)
                    if matches and matches[0][1] >= 80:
                        suggested_word = matches[0][0]
                        if original_word.lower() != suggested_word and not re.search(r'\d', original_word):
                            # Need to prompt user
                            session['original_word'] = original_word
                            session['suggested_word'] = suggested_word
                            # Prompt the user again
                            history.append(f"Potential misspelling detected: '{session['original_word']}' in '{session['current_name']}'")
                            history.append(f"Would you like to update it to '{session['suggested_word']}'? (y/n):")
                            return render_template('terminal.html', history=history)
                        else:
                            corrected_words.append(original_word)
                    else:
                        corrected_words.append(original_word)

            # Generate the corrected name and alias after processing all words
            processed_name = ''.join(corrected_words)
            alias_name = re.sub(r'\([^)]*\)', '', processed_name)
            alias_name = re.sub(r"['+/,<>:\.\-]", ' ', alias_name)
            alias_name = alias_name.replace('%', ' PCT')
            alias_words = re.findall(r'\b[\w*-]+\b', alias_name)
            alias_corrected_words = [process_alias_word(word, custom_dict) for word in alias_words]
            processed_alias = '_'.join(filter(bool, alias_corrected_words))
            if session['append_text'].upper() not in processed_alias:
                processed_alias = f"{processed_alias.rstrip('_')}_{session['append_text'].upper()}"

            # Append to converted names
            converted_names = session.get('converted_names', [])
            converted_aliases = session.get('converted_aliases', [])
            converted_names.append(processed_name)
            converted_aliases.append(processed_alias)
            session['converted_names'] = converted_names
            session['converted_aliases'] = converted_aliases

            # Reset corrected words and continue processing
            session['corrected_words'] = []
            session['state'] = 'processing_multi_add'
            # Call the function again to process the next name
            return alias_converter()
        else:
            session.clear()
            history = []
            history.append("Do you want to use spell check? (y/n):")
            session['state'] = 'start'
            session['history'] = history

        # Update the session
        session['history'] = history
    else:
        # Handle GET requests without resetting the application
        history = session.get('history', [])
        return render_template('terminal.html', history=history)

    return render_template('terminal.html', history=history)

# Code Compare Dropped and Duplicates Report route
@app.route('/compare-dropped-duplicates', methods=['GET', 'POST'])
def compare_dropped_duplicates():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        if uploaded_file and allowed_file(uploaded_file.filename):
            filename = secure_filename(uploaded_file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(file_path)

            # Call the function to process the file (rename columns + dropped/duplicates)
            processed_file_path = process_dropped_duplicates(file_path)

            # Return the processed file for download
            return send_file(processed_file_path, as_attachment=True)

    return render_template('compare_dropped_duplicates.html')

# Rename columns and run dropped/duplicates logic
def process_dropped_duplicates(file_path):
    # Step 1: Load the Excel file and rename columns
    df = pd.read_excel(file_path)
    df.rename(columns={df.columns[3]: 'Old Concept Name', df.columns[6]: 'New Concept Name'}, inplace=True)

    # Output columns
    columns_to_output = ['Code System', 'Code', 'Names', 'Old Concept Name', 'New Concept Name']

    # Function Definitions (from your original script)
    # Function - All Duplicates (Red, White, and Green)
    def process_all_dups(df):
        df['Normalized_Code'] = df['Code'].astype(str).str.rstrip('.0')
        grouped = df.groupby(['Code System', 'Normalized_Code'])
        matching_codes = []
        for (code_system, code), group in grouped:
            removed_from_concept = group['Old Concept Name'].notnull() & group['New Concept Name'].isnull()
            added_to_concept = group['Old Concept Name'].isnull() & group['New Concept Name'].notnull()
            remains_in_concept = group['Old Concept Name'].notnull() & group['New Concept Name'].notnull()
            if removed_from_concept.any() and added_to_concept.any() and remains_in_concept.any():
                matching_codes.append((code_system, code))
        matching_codes_df = df[(df[['Code System', 'Normalized_Code']].apply(tuple, axis=1).isin(matching_codes))]
        return matching_codes_df[columns_to_output]

    # Function - Green/White Duplicates
    def process_gw_dups(df):
        removed_codes = df[df['Old Concept Name'].notnull() & df['New Concept Name'].isnull()]['Code'].unique()
        initial_matches = df[df['Old Concept Name'].notnull() & df['New Concept Name'].notnull()]
        duplicates = []
        for _, group in initial_matches.groupby(['Code System', 'Code']):
            initial_match_rows = group
            subsequent_rows = df[(df['Code System'] == group['Code System'].iloc[0]) &
                                (df['Code'] == group['Code'].iloc[0]) &
                                (df.index > initial_match_rows.index.max())]
            if (subsequent_rows['Old Concept Name'].isnull() & subsequent_rows['New Concept Name'].notnull()).any():
                duplicates.append((group['Code System'].iloc[0], group['Code'].iloc[0]))
        duplicates = [dup for dup in duplicates if dup[1] not in removed_codes]
        duplicate_codes_df = df[(df[['Code System', 'Code']].apply(tuple, axis=1).isin(duplicates))]
        return duplicate_codes_df[duplicate_codes_df['New Concept Name'].notnull()][columns_to_output]

    # Function - Red/White Duplicates
    def process_rw_dups(df):
        df['Code'] = df['Code'].astype(str)
        added_codes = df[df['Old Concept Name'].isnull() & df['New Concept Name'].notnull()]['Code'].unique()
        df = df[~df['Code'].isin(added_codes)]
        initial_matches = df[df['Old Concept Name'].notnull() & df['New Concept Name'].notnull()]
        matching_codes = []
        for _, group in initial_matches.groupby(['Code System', 'Code']):
            subsequent_rows = df[(df['Code System'] == group['Code System'].iloc[0]) & (df['Code'] == group['Code'].iloc[0])]
            if (subsequent_rows['Old Concept Name'].notnull() & subsequent_rows['New Concept Name'].isnull()).any():
                matching_codes.append((group['Code System'].iloc[0], group['Code'].iloc[0]))
        matching_codes_df = df[df[['Code System', 'Code']].apply(tuple, axis=1).isin(matching_codes)]
        return matching_codes_df[columns_to_output]
    
    # Function - Dropped
    def process_dropped_codes(df):
        removed_not_added = []
        for code in df['Code'].unique():
            code_df = df[df['Code'] == code]
            if code_df['Old Concept Name'].notnull().any() and not code_df['New Concept Name'].notnull().any():
                removed_not_added.append(code)
        removed_not_added_df = df[df['Code'].isin(removed_not_added)]
        columns_to_output_dropped = ['Code System', 'Code', 'Names', 'Old Concept Name']
        return removed_not_added_df[columns_to_output_dropped]

    # Function - Code Additions
    def process_code_additions(df):
        added_codes_df = df[df['Old Concept Name'].isnull() & df['New Concept Name'].notnull()]
        columns_to_output_additions = ['Code System', 'Code', 'Names', 'New Concept Name']
        return added_codes_df[columns_to_output_additions]

    # Run the processing functions
    all_dups_df = process_all_dups(df)
    gw_dups_df = process_gw_dups(df)
    rw_dups_df = process_rw_dups(df)
    dropped_codes_df = process_dropped_codes(df)
    code_additions_df = process_code_additions(df)

    # Load the workbook (use openpyxl)
    wb = load_workbook(file_path)

    # Create Dropped sheet
    if "Dropped" in wb.sheetnames:
        wb.remove(wb["Dropped"])
    ws_dropped = wb.create_sheet("Dropped", 1)

    # Dropped codes data and styles
    for col, column_name in enumerate(dropped_codes_df.columns, start=1):
        ws_dropped.cell(row=1, column=col, value=column_name)
        ws_dropped.cell(row=1, column=col).font = Font(bold=True)
    for row_idx, data_row in enumerate(dropped_codes_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(data_row, start=1):
            ws_dropped.cell(row=row_idx, column=col_idx, value=value)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color='9C0006')
    for row in ws_dropped.iter_rows(min_row=2, max_row=ws_dropped.max_row, min_col=1, max_col=ws_dropped.max_column):
        old_concept = row[3].value
        if old_concept:
            for cell in row:
                cell.fill = red_fill
                cell.font = red_font

    # Create Duplicates sheet
    if "Duplicates" in wb.sheetnames:
        wb.remove(wb["Duplicates"])
    ws = wb.create_sheet("Duplicates")

    # Write data for Duplicates sheet with headers
    def write_data_with_header(ws, start_row, header, data):
        ws.cell(row=start_row, column=1, value=header)
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(columns_to_output))
        
        for col, column_name in enumerate(columns_to_output, start=1):
            ws.cell(row=start_row + 1, column=col, value=column_name)
            ws.cell(row=start_row + 1, column=col).font = Font(bold=True)
        
        for row_idx, data_row in enumerate(data.itertuples(index=False), start=start_row + 2):
            for col_idx, value in enumerate(data_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        return start_row + data.shape[0] + 3 

    # Duplicates sheet data and styles
    next_row = write_data_with_header(ws, 1, "Green, White, and Red Duplicates", all_dups_df)
    next_row = write_data_with_header(ws, next_row, "Green and White Duplicates", gw_dups_df)
    next_row = write_data_with_header(ws, next_row, "Red and White Duplicates", rw_dups_df)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color='006100')

    # Dictionary to store the rows for each unique code
    code_system_rows = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        code_system = row[0].value  # Column A
        code = row[1].value  # Column B
        old_concept = row[3].value
        new_concept = row[4].value
        
        if isinstance(old_concept, str) and not isinstance(new_concept, str):
            for cell in row:
                cell.fill = red_fill
                cell.font = red_font
        elif not isinstance(old_concept, str) and isinstance(new_concept, str):
            for cell in row:
                cell.fill = green_fill
                cell.font = green_font
        
        # Store the row number for each unique code and code system combination
        key = (code_system, code)
        if key not in code_system_rows:
            code_system_rows[key] = []
        code_system_rows[key].append(row[0].row)

    # Add borders to the outer edges of each set of codes
    medium_side = Side(style='medium')

    for (code_system, code), rows in code_system_rows.items():
        if len(rows) > 1:  # Only add borders if there's more than one row with the same code and code system
            min_row = min(rows)
            max_row = max(rows)
            
            # Apply top border
            for cell in ws[min_row]:
                cell.border = cell.border + Border(top=medium_side)
            
            # Apply bottom border
            for cell in ws[max_row]:
                cell.border = cell.border + Border(bottom=medium_side)
            
            # Apply left and right borders
            for row_idx in range(min_row, max_row + 1):
                ws.cell(row=row_idx, column=1).border = ws.cell(row=row_idx, column=1).border + Border(left=medium_side)
                ws.cell(row=row_idx, column=ws.max_column).border = ws.cell(row=row_idx, column=ws.max_column).border + Border(right=medium_side)

    # Create Code Additions sheet and styles
    if "Additions" in wb.sheetnames:
        wb.remove(wb["Additions"])
    ws_added = wb.create_sheet("Additions", 2)
    for col, column_name in enumerate(code_additions_df.columns, start=1):
        ws_added.cell(row=1, column=col, value=column_name)
        ws_added.cell(row=1, column=col).font = Font(bold=True)
    for row_idx, data_row in enumerate(code_additions_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(data_row, start=1):
            ws_added.cell(row=row_idx, column=col_idx, value=value)

    for row in ws_added.iter_rows(min_row=2, max_row=ws_added.max_row, min_col=1, max_col=ws_added.max_column):
        for cell in row:
            cell.fill = green_fill
            cell.font = green_font

    # Save the workbook to the processed files folder
    output_file_name = 'processed_' + os.path.basename(file_path)
    output_file_path = os.path.join(app.config['PROCESSED_FOLDER'], output_file_name)
    wb.save(output_file_path)

    return output_file_path

# Code Compare Grouper Check route
@app.route('/compare-grouper-check', methods=['GET', 'POST'])
def compare_grouper_check():
    if request.method == 'POST':
        # Handle file upload
        uploaded_file = request.files.get('file')
        if uploaded_file and allowed_file(uploaded_file.filename):
            filename = secure_filename(uploaded_file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(file_path)

            # Process the uploaded file
            try:
                processed_file_path = process_grouper_check(file_path)

                # Return the processed file for download
                return send_file(processed_file_path, as_attachment=True)
            except Exception as e:
                # Log the error and return an error message
                print(f"Error processing file: {e}")
                error_message = f"An error occurred during processing: {e}"
                return render_template('compare_grouper_check.html', error=error_message)
        else:
            # File not allowed or not provided
            error_message = "Please upload a valid .xlsx file."
            return render_template('compare_grouper_check.html', error=error_message)

    elif request.method == 'GET' and 'download_template' in request.args:
        # User requested to download the template
        template_file_path = generate_template()
        return send_file(template_file_path, as_attachment=True)

    return render_template('compare_grouper_check.html')

def generate_template():
    # Create a new workbook
    wb = Workbook()

    # Create the Instructions sheet
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"

    # List of instruction lines
    instructions_lines = [
        "Concepts Tab - add concept names from the more recent release sorted alphabetically",
        "Relationships tab - add concept names from Concept 2 column sorted alphabetically",
        "Note - please do not apply any formatting or links"
    ]

    # Write each instruction line into separate rows
    for idx, line in enumerate(instructions_lines, start=1):
        ws_instructions.cell(row=idx, column=1, value=line)

    # Create the Concepts Tab
    ws_concepts = wb.create_sheet("Concepts Tab")
    ws_concepts.cell(row=1, column=1, value="Concept Name")

    # Create the Relationships Tab
    ws_relationships = wb.create_sheet("Relationships Tab")
    ws_relationships.cell(row=1, column=1, value="Concept Name")

    # Save the workbook to a temporary file
    template_file_path = os.path.join(app.config['PROCESSED_FOLDER'], 'Groupers_Template.xlsx')
    wb.save(template_file_path)

    return template_file_path

# Placeholder for Grouper Check processing
def process_grouper_check(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Get the sheets
    try:
        concepts_sheet = wb["Concepts Tab"]
        relationships_sheet = wb["Relationships Tab"]
    except KeyError as e:
        raise Exception(f"Missing sheet in the uploaded file: {e}")

    # Extract values from both sheets (skip the header)
    concepts_values = set()
    for row in concepts_sheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell:
                concepts_values.add(cell)

    relationships_values = set()
    for row in relationships_sheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell:
                relationships_values.add(cell)

    # Find unique and common values
    unique_to_concepts = concepts_values - relationships_values
    unique_to_relationships = relationships_values - concepts_values
    common_values = concepts_values.intersection(relationships_values)

    # Create a new sheet for results
    if "Check Groupers" in wb.sheetnames:
        wb.remove(wb["Check Groupers"])
    check_groupers = wb.create_sheet("Check Groupers")

    # Write headers and unique values
    headers = ["Unique to Concepts Tab", "Unique to Relationships Tab", "No Review Required"]
    row = 1
    for header in headers:
        check_groupers.cell(row=row, column=1, value=header).font = Font(bold=True)
        row += 1

        if header == "Unique to Concepts Tab":
            for value in sorted(unique_to_concepts):
                check_groupers.cell(row=row, column=1, value=value)
                row += 1
        elif header == "Unique to Relationships Tab":
            for value in sorted(unique_to_relationships):
                check_groupers.cell(row=row, column=1, value=value)
                row += 1
        else:  # No Review Required
            check_groupers.cell(row=row, column=1, value="Concepts Tab")
            check_groupers.cell(row=row, column=2, value="Relationships Tab")
            row += 1
            for value in sorted(common_values):
                check_groupers.cell(row=row, column=1, value=value)
                check_groupers.cell(row=row, column=2, value=value)
                row += 1

        row += 1  # Add a blank row after each section

    # Auto-adjust column widths
    for column_cells in check_groupers.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        check_groupers.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Save the workbook to the processed files folder
    output_file_name = 'processed_grouper_' + os.path.basename(file_path)
    output_file_path = os.path.join(app.config['PROCESSED_FOLDER'], output_file_name)
    wb.save(output_file_path)

    return output_file_path

def analyze_excel(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Get the sheets
    sheet1 = wb['Sheet1']
    sheet2 = wb['Sheet2']
    # Remove 'Sheet3' if it already exists to avoid duplicate sheets
    if 'Sheet3' in wb.sheetnames:
        wb.remove(wb['Sheet3'])
    sheet3 = wb.create_sheet('Sheet3')
    
    # Function to clean concept names
    def clean_concept(concept):
        if not concept:
            return ''
        # Remove text within parentheses
        concept_clean = re.sub(r'\(.*?\)', '', concept)
        # Strip and lowercase
        concept_clean = concept_clean.strip().lower()
        return concept_clean
    
    # Extract and clean concepts from Sheet1
    concepts = set()
    for row in sheet1.iter_rows(min_row=2, values_only=True):
        if row[0]:
            concept = clean_concept(row[0])
            if concept:  # Only add non-empty concepts
                concepts.add(concept)
    
    # Write header to Sheet3
    sheet3.append(['Code System Name', 'Code', 'Names', 'Concept Names'])
    
    # Filter and write data to Sheet3
    for row in sheet2.iter_rows(min_row=2, values_only=True):
        if len(row) < 4 or not all(row[:4]):
            # Skip rows that don't have all required columns
            continue
        
        code_system, code, names, concept_names = row
        
        # Split concept names on newlines only, and remove empty strings
        row_concepts = [concept.strip() for concept in re.split(r'[\n]+', concept_names) if concept.strip()]
        # Clean concepts
        row_concepts_clean = [clean_concept(concept) for concept in row_concepts]
        
        match_found = False
        for concept_clean in row_concepts_clean:
            if concept_clean and concept_clean in concepts:
                match_found = True
                break
        
        if match_found:
            sheet3.append([code_system, code, names, concept_names])
    
    # Save the workbook to the processed files folder
    output_file_name = 'processed_' + os.path.basename(file_path)
    output_file_path = os.path.join(app.config['PROCESSED_FOLDER'], output_file_name)
    wb.save(output_file_path)
    
    return output_file_path

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}

@app.route('/non-exclusive-personalized', methods=['GET', 'POST'])
def non_exclusive_personalized():
    if request.method == 'POST':
        uploaded_file = request.files.get('file')
        if uploaded_file and allowed_file(uploaded_file.filename):
            filename = secure_filename(uploaded_file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(file_path)
            
            try:
                processed_file_path = analyze_excel(file_path)
                return send_file(processed_file_path, as_attachment=True)
            except Exception as e:
                error_message = f"An error occurred during processing: {e}"
                return render_template('nonexclusive_personalized.html', error=error_message)
        else:
            error_message = "Please upload a valid .xlsx file."
            return render_template('nonexclusive_personalized.html', error=error_message)
    return render_template('nonexclusive_personalized.html') 

def get_snomed_words():
    # Use the existing database connection
    cursor = db.cursor()
    
    # Query the active descriptions from the database
    cursor.execute("SELECT term FROM Descriptions WHERE active=1")
    snomed_terms = cursor.fetchall()
    cursor.close()
    
    # Extract words and convert them to a set
    words = set()
    for term in snomed_terms:
        if term[0]:
            words.update(re.findall(r'\b\w+\b', term[0].lower()))
    return words

snomed_words = get_snomed_words()
print(f"Loaded {len(snomed_words)} unique words from SNOMED CT")

@app.route('/snomed_spell_checker', methods=['GET', 'POST'])
def snomed_spell_checker():
    if request.method == 'POST':
        # Handle file upload
        if 'file' not in request.files:
            return "No file uploaded", 400
        file = request.files['file']
        if file.filename == '':
            return "No file selected", 400
        if file:
            # Save the uploaded file to a temporary location
            filename = secure_filename(file.filename)
            temp_dir = tempfile.gettempdir()
            input_file_path = os.path.join(temp_dir, filename)
            file.save(input_file_path)
            
            # Run the script on the uploaded file
            try:
                output_file_path = run_spell_check_script(input_file_path)
                # After processing, send the output file to the user
                return send_file(output_file_path, as_attachment=True)
            except Exception as e:
                return f"An error occurred: {e}", 500
    else:
        return render_template('snomed_spell_checker.html')

def run_spell_check_script(input_file_path):
    # Read the input Excel file
    df = pd.read_excel(input_file_path)

    # Standardize column names
    df.rename(columns={df.columns[3]: 'Old Concept Name', df.columns[6]: 'New Concept Name'}, inplace=True)

    # Run spell check using the preloaded SNOMED words
    spell_check_results = perform_spell_check(df, 'New Concept Name', snomed_words)

    # Save the results to a new Excel file
    output_file_path = os.path.splitext(input_file_path)[0] + '_spell_check.xlsx'

    # Write results to Excel
    wb = load_workbook(input_file_path)
    if "Spell Check Report" in wb.sheetnames:
        wb.remove(wb["Spell Check Report"])
    ws_spell_check = wb.create_sheet("Spell Check Report")
    for col, column_name in enumerate(spell_check_results.columns, start=1):
        ws_spell_check.cell(row=1, column=col, value=column_name)
        ws_spell_check.cell(row=1, column=col).font = Font(bold=True)
    for row, data_row in enumerate(spell_check_results.itertuples(index=False), start=2):
        for col, value in enumerate(data_row, start=1):
            ws_spell_check.cell(row=row, column=col, value=value)

    wb.save(output_file_path)

    return output_file_path

def perform_spell_check(df, column_name, snomed_words):
    spell_check_results = []

    unique_concepts = df[column_name].dropna().unique()

    for concept_name in unique_concepts:
        misspelled_words = []
        words = re.findall(r'\b\w+\b', concept_name.lower())
        
        for word in words:
            if word not in snomed_words:
                # Check if it's a potential misspelling
                matches = process.extract(word, snomed_words, limit=1, scorer=fuzz.ratio)
                if matches and matches[0][1] >= 80:  # If similarity is 80% or higher
                    misspelled_words.append(f"{word} (possible: {matches[0][0]})")
                else:
                    misspelled_words.append(word)
        
        if misspelled_words:
            description = f"Potential misspellings: {', '.join(misspelled_words)}"
            needs_review = 'Yes'
        else:
            description = "No spelling issues detected"
            needs_review = 'No'

        spell_check_results.append({
            'Original Value': concept_name,
            'Description': description,
            'Needs Review': needs_review
        })

    return pd.DataFrame(spell_check_results)

ALLOWED_EXTENSIONS = {'xlsx'}

class ProgressTracker:
    def __init__(self):
        self.progress = 0
        self.message = ""
        
    def update(self, progress, message=""):
        self.progress = progress
        self.message = message

progress_tracker = ProgressTracker()

@app.route('/progress')
def progress():
    def generate():
        while True:
            # Send progress updates
            data = {
                'progress': progress_tracker.progress,
                'message': progress_tracker.message
            }
            yield f"data: {json.dumps(data)}\n\n"
            time.sleep(0.5)  # Update every 500ms
            if progress_tracker.progress >= 100:
                break
    return Response(stream_with_context(generate()), mimetype='text/event-stream')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def standardize_column_names(file_path: str) -> str:
    """
    Standardizes column names by renaming specific columns to 'Old Concept Name' and 'New Concept Name'.
    Returns the path of the updated file.
    """
    # Load the Excel file
    df = pd.read_excel(file_path)
    
    # Rename specific columns
    if len(df.columns) > 6:  # Ensure there are enough columns to rename
        df.rename(columns={df.columns[3]: 'Old Concept Name', df.columns[6]: 'New Concept Name'}, inplace=True)
    
    # Save the modified file back to the same path
    df.to_excel(file_path, index=False)
    return file_path

def get_sheet_name(file_path: str) -> str:
    """Get the name of the first sheet in the workbook."""
    try:
        xlsx = openpyxl.load_workbook(file_path, read_only=True)
        sheet_name = xlsx.sheetnames[0]  # Get the first sheet name
        xlsx.close()
        return sheet_name
    finally:
        del xlsx
        gc.collect()

#code additions summary
def code_additions(file_path: str, sheet_name: str, generative_ai_inference_client: GenerativeAiInferenceClient, compartment_id: str) -> str:
    output_file = os.path.join(os.path.dirname(file_path), 'additions_output.txt')
    flagged_file = os.path.join(os.path.dirname(file_path), 'additions_flagged.csv')
    
    # Call the function with all required arguments
    process_large_dataset_stream_additions(
        file_path=file_path,
        sheet_name=sheet_name,
        output_file=output_file,
        flagged_file=flagged_file,
        generative_ai_inference_client=generative_ai_inference_client,
        compartment_id=compartment_id
    )
    
    return output_file

# Reduce logging level for Cohere library
logging.getLogger('cohere').setLevel(logging.WARNING)

EXCLUDED_CODE_SYSTEMS = {'Read Codes v2', 'Read Codes v3', 'ICD-10-SE'}

def clean_text_additions(text: str) -> str:
    if text is None:
        return ""
    if not isinstance(text, str):
        try:
            text = str(text)
        except:
            return ""
    # Preserve alphanumeric characters, spaces, hyphens, commas, and periods
    text = re.sub(r'[^\w\s\-,.]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extract_active_code_additions(text: str) -> str:
    if not isinstance(text, str):
        return ""

    lines = text.strip().split('\n')

    asterisk_match = re.findall(r'\*(.*?)\*', text)
    if asterisk_match:
        return asterisk_match[0].strip()

    for line in reversed(lines):
        parentheses_match = re.match(r'(.*)\s+\([^)]+\)$', line)
        if parentheses_match:
            return parentheses_match.group(1).strip()

    return lines[-1].strip()

def stream_excel_additions(file_path: str, sheet_name: str = None, chunk_size: int = 1000) -> Iterator[pd.DataFrame]:
    try:
        xlsx = openpyxl.load_workbook(file_path, read_only=True)
        if sheet_name is None:
            sheet_name = xlsx.sheetnames[0]  # Use first sheet if none specified
        sheet = xlsx[sheet_name]
        
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)  # Get the header row
        
        chunk = []
        for row in rows:
            chunk.append(row)
            if len(chunk) == chunk_size:
                df = pd.DataFrame(chunk, columns=header)
                yield df[['Code System', 'Code', 'Names', 'Old Concept Name', 'New Concept Name']]
                chunk = []
        
        if chunk:  # Don't forget the last chunk if it's not full
            df = pd.DataFrame(chunk, columns=header)
            yield df[['Code System', 'Code', 'Names', 'Old Concept Name', 'New Concept Name']]
    finally:
        xlsx.close()
        del xlsx
        gc.collect()   

def preprocess_row_additions(row: pd.Series) -> Tuple[str, str, str, str]:
    code_system = clean_text_additions(str(row.get('Code System', '')))
    code = clean_text_additions(str(row.get('Code', '')))
    name = clean_text_additions(extract_active_code_additions(str(row.get('Names', ''))))
    new_concept = clean_text_additions(str(row.get('New Concept Name', '')))
    return code_system, code, name, new_concept

def validate_new_concept_additions(new_concept: str, name: str) -> Tuple[bool, str]:
    """
    Validate the New Concept Name entry and return a tuple of (is_valid, reason).
    """
    if not new_concept or new_concept.isspace():
        return False, "Empty or whitespace"
    
    if len(new_concept) < 3:
        return False, "Unusually short"
    
    if len(new_concept) > 200:
        return False, "Unusually long"
    
    common_placeholders = ['N/A', 'None', 'Unknown', 'TBD', 'To Be Determined']
    if new_concept.lower() in (placeholder.lower() for placeholder in common_placeholders):
        return False, f"Common placeholder: {new_concept}"
    
    if new_concept.lower() == name.lower():
        return False, "Identical to Names entry"
    
    return True, "Valid"

def process_chunk_additions(chunk: pd.DataFrame) -> Tuple[Dict[str, Set[str]], List[Tuple[str, str, str, str]]]:
    grouped = defaultdict(set)
    flagged_entries = []
    
    for _, row in chunk.iterrows():
        code_system, code, name, new_concept = preprocess_row_additions(row)
        old_concept = clean_text_additions(str(row.get('Old Concept Name', '')))
        
        if code_system not in EXCLUDED_CODE_SYSTEMS:
            is_valid, reason = validate_new_concept_additions(new_concept, name)
            if is_valid and new_concept != old_concept:
                grouped[new_concept].add(name)
                if old_concept:
                    grouped[old_concept].discard(name)
            elif not is_valid:
                flagged_entries.append((new_concept, code_system, code, reason))
    
    # Remove any concepts that ended up with no codes
    grouped = {k: v for k, v in grouped.items() if v}
    
    return grouped, flagged_entries

def retry_on_throttle(func: Callable) -> Callable:
    """Decorator to handle OCI throttling with exponential backoff."""
    @retry(
        retry=retry_if_exception_type(oci.exceptions.ServiceError),
        wait=wait_exponential(multiplier=1, min=4, max=10),
        stop=stop_after_attempt(3),
        before_sleep=lambda retry_state: logging.info(f"Retry attempt {retry_state.attempt_number}, waiting {retry_state.next_action.sleep} seconds...")
    )
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except oci.exceptions.ServiceError as e:
            if e.status == 429:  # Too Many Requests
                logging.warning(f"Rate limit hit, retrying with backoff: {e}")
                raise  # Let tenacity handle the retry
            raise  # Re-raise other service errors
    return wrapper

def summarize_concept_additions_with_semaphore(*args, **kwargs):
    with semaphore:
        return summarize_concept_additions(*args, **kwargs)

def summarize_concept_additions(concept: str, names: Set[str], generative_ai_inference_client: GenerativeAiInferenceClient, compartment_id: str) -> str:
    if not names:
        return f"{concept}: No new codes were added to this concept."

    names_text = "\n".join(names)

    prompt = f"""Summarize the following information for the concept '{concept}':

{names_text}
        "1. Begin by listing out every unique value in the 'New Concept Name' column. Every unique concept should be listed only once, with no duplicates or omissions.\n"
        "2. For each concept identified in step 1, concisely summarize the values in the 'Names' column that are associated with that Concept.\n"
        "3. Present the summary in the following format: Concept: [preamble] for [code summaries]. Rotate through the following preambles for variety: 'Included codes', 'Added codes specifying', 'Codes were added', 'Added codes for'.\n"
        "4. Do not interpret any of the codes. Only provide a summarization of them.\n"
        "5. Do not include any references to 'Code System' or 'Code' in your summary\n"
        "Example Input:\n"
        "Concepts, Codes\n"
        "Colon Cancer, Stage I Colon Cancer AJCC v8\n"
        "Colon Cancer, Stage I Colon Cancer AJCC v6 and v7\n"
        "Colon Cancer, Ca transverse colon\n"
        "Colon Cancer, Ca descending colon\n"
        "Colon Cancer, Ca ascending colon\n"
        "Colon Cancer, Ca splenic flexure - colon\n"
        "Colon Cancer, Ca sigmoid colon\n"
        "Desired AI Output:\n"
        "Colon Cancer: Includes codes related to colon cancer of various regions of the colon, including transverse, ascending, descending, and sigmoid colon and splenic flexure. Additionally, codes related to staging by various systems, such as AJCC (American Joint Committee on Cancer)."
    )
Format the summary as:
{concept}: [Your summary here]"""

    chat_request = CohereChatRequest(
        message=prompt,
        max_tokens=600,
        temperature=0.25,
        frequency_penalty=0,
        top_p=0.75,
        top_k=0
    )

    chat_detail = ChatDetails(
        serving_mode=OnDemandServingMode(
            model_id="ocid1.generativeaimodel.oc1.us-chicago-1.amaaaaaask7dceya7ozidbukxwtun4ocm4ngco2jukoaht5mygpgr6gq2lgq"
        ),
        chat_request=chat_request,
        compartment_id=compartment_id
    )

    @retry_on_throttle
    def make_chat_request(chat_detail):
        return generative_ai_inference_client.chat(chat_detail)

    try:
        chat_response = make_chat_request(chat_detail)
        
        # Convert response to string and parse as JSON, similar to the working async version
        response_json = json.loads(str(chat_response.data))
        
        # Extract the text from the parsed JSON
        summary_text = response_json.get('chat_response', {}).get('text', '')
        
        if summary_text:
            return summary_text.strip()
        else:
            logging.warning(f"No 'text' field found in the response for concept '{concept}'. Full response: {response_json}")
            return f"{concept}: Unable to generate summary."

    except json.JSONDecodeError as e:
        logging.error(f"Error parsing JSON response for concept '{concept}': {e}")
        return f"{concept}: Error occurred while parsing the response."
    except oci.exceptions.ServiceError as e:
        if e.status == 429:
            logging.error(f"Rate limit exceeded even after retries for concept '{concept}': {e}")
        else:
            logging.error(f"Service error for concept '{concept}': {e}")
        return f"{concept}: Error occurred while summarizing this concept."
    except Exception as e:
        logging.error(f"Unexpected error summarizing concept '{concept}': {str(e)}")
        return f"{concept}: Unexpected error occurred while summarizing this concept."
    
def get_total_rows_additions(file_path: str) -> int:
    try:
        xl = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = xl.active  # Selects the first sheet in the workbook
        total_rows = sheet.max_row - 1  # Subtract 1 to exclude header
        return total_rows
    finally:
        xl.close()
        del xl
        gc.collect()

def write_report_additions(output_file: str, summaries: List[str], deprecated_concepts: Set[str]) -> None:
    # Sort summaries alphabetically by concept name
    sorted_summaries = sorted(summaries, key=lambda x: x.split(':', 1)[0].lower())
    
    with open(output_file, "w", encoding='utf-8') as f:
        # Write introductory sentence
        f.write("The following concepts were created or modified and codes were added as described:\n\n")
        
        # Write summaries
        for summary in sorted_summaries:
            f.write(f"{summary}\n\n")
        
        # Write deprecated concepts
        if deprecated_concepts:
            f.write("\nThe following concepts were deprecated:\n")
            for concept in sorted(deprecated_concepts):
                f.write(f"{concept}\n")

def process_large_dataset_stream_additions(
    file_path: str,
    sheet_name: str,
    output_file: str,
    flagged_file: str,
    generative_ai_inference_client: GenerativeAiInferenceClient,
    compartment_id: str
) -> None:
    sheet_name = get_sheet_name(file_path)
    total_rows = get_total_rows_additions(file_path)
    data_stream = stream_excel_additions(file_path, sheet_name)

    all_data = defaultdict(set)
    all_flagged_entries = []
    deprecated_concepts = set()
    processed_rows = 0

    # Process chunks with progress updates
    with tqdm(total=total_rows, desc="Processing", unit="rows") as pbar:
        for chunk in data_stream:
            chunk_data, flagged_entries = process_chunk_additions(chunk)
            for concept, names in chunk_data.items():
                all_data[concept].update(names)
            all_flagged_entries.extend(flagged_entries)

            chunk_size = len(chunk)
            processed_rows += chunk_size
            pbar.update(chunk_size)

            # Update progress tracker
            progress = min(95, int((processed_rows / total_rows) * 100))
            progress_tracker.update(progress, f"Processing rows: {processed_rows}/{total_rows}")

    sorted_concepts = dict(sorted(all_data.items()))

    # Generate summaries with semaphore-controlled concurrency
    progress_tracker.update(96, "Generating summaries...")
    summaries = []

    concepts_list = list(sorted_concepts.items())

    for i in range(0, len(concepts_list), batch_size):
        batch = concepts_list[i:i + batch_size]
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_concept = {
                executor.submit(
                    summarize_concept_additions_with_semaphore,  # Use the wrapper function
                    concept,
                    names,
                    generative_ai_inference_client,
                    compartment_id
                ): concept
                for concept, names in batch
            }

            for future in as_completed(future_to_concept):
                concept = future_to_concept[future]
                try:
                    summary = future.result()
                    summaries.append(summary)
                except Exception as e:
                    logging.error(f"Error processing concept {concept}: {str(e)}")
                    summaries.append(f"{concept}: Error generating summary")

    progress_tracker.update(98, "Writing report...")
    write_report_additions(output_file, summaries, deprecated_concepts)

    with open(flagged_file, "w", encoding='utf-8') as f:
        f.write("Concept Name,Code System,Code,Reason\n")
        for entry in sorted(all_flagged_entries):
            f.write(f"{','.join(entry)}\n")

    progress_tracker.update(100, "Processing complete")

#code removals summary
def code_removals(file_path: str, sheet_name: str, generative_ai_inference_client: GenerativeAiInferenceClient, compartment_id: str) -> str:
    output_file = os.path.join(os.path.dirname(file_path), 'removals_output.txt')
    flagged_file = os.path.join(os.path.dirname(file_path), 'removals_flagged.csv')

    try:
        # Call the function with all required arguments
        process_large_dataset_stream_removals(
            file_path=file_path,
            sheet_name=sheet_name,
            output_file=output_file,
            flagged_file=flagged_file,
            generative_ai_inference_client=generative_ai_inference_client,
            compartment_id=compartment_id
        )
    except Exception as e:
        logging.error(f"An error occurred during processing: {e}")
        raise

    if not os.path.exists(output_file):
        raise FileNotFoundError(f"Expected output file not found: {output_file}")

    return output_file

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Set up HTTP request logging
http_client.HTTPConnection.debuglevel = 0  # Change to 1 for more verbose output

logging.getLogger("requests").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)

EXCLUDED_CODE_SYSTEMS = {'Read Codes v2', 'Read Codes v3', 'ICD-10-SE'}

def clean_text_removals(text: str) -> str:
    if text is None:
        return ""
    if not isinstance(text, str):
        try:
            text = str(text)
        except:
            return ""
    # Preserve alphanumeric characters, spaces, hyphens, commas, and periods
    text = re.sub(r'[^\w\s\-,.]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extract_active_code_removals(text: str) -> str:
    if not isinstance(text, str):
        return ""

    lines = text.strip().split('\n')

    asterisk_match = re.findall(r'\*(.*?)\*', text)
    if asterisk_match:
        return asterisk_match[0].strip()

    for line in reversed(lines):
        parentheses_match = re.match(r'(.*)\s+\([^)]+\)$', line)
        if parentheses_match:
            return parentheses_match.group(1).strip()

    return lines[-1].strip()

def stream_excel_removals(file_path: str, sheet_name: str = None, chunk_size: int = 1000) -> Iterator[pd.DataFrame]:
    try:
        xlsx = openpyxl.load_workbook(file_path, read_only=True)
        if sheet_name is None:
            sheet_name = xlsx.sheetnames[0]  # Use first sheet if none specified
        sheet = xlsx[sheet_name]
        
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)  # Get the header row
        
        chunk = []
        for row in rows:
            chunk.append(row)
            if len(chunk) == chunk_size:
                df = pd.DataFrame(chunk, columns=header)
                yield df[['Code System', 'Code', 'Names', 'Old Concept Name', 'New Concept Name']]
                chunk = []
        
        if chunk:  # Don't forget the last chunk if it's not full
            df = pd.DataFrame(chunk, columns=header)
            yield df[['Code System', 'Code', 'Names', 'Old Concept Name', 'New Concept Name']]
    finally:
        xlsx.close()
        del xlsx
        gc.collect()

def preprocess_row_removals(row: pd.Series) -> Tuple[str, str, str, str, str]:
    code_system = clean_text_removals(str(row.get('Code System', '')))
    code = clean_text_removals(str(row.get('Code', '')))
    name = clean_text_removals(extract_active_code_removals(str(row.get('Names', ''))))
    old_concept = clean_text_removals(str(row.get('Old Concept Name', '')))
    new_concept = clean_text_removals(str(row.get('New Concept Name', '')))
    return code_system, code, name, old_concept, new_concept

def validate_old_concept_removals(old_concept: str, name: str) -> Tuple[bool, str]:
    """
    Validate the Old Concept Name entry and return a tuple of (is_valid, reason).
    """
    if not old_concept or old_concept.isspace():
        return False, "Empty or whitespace"
    
    if len(old_concept) < 3:
        return False, "Unusually short"
    
    if len(old_concept) > 200:
        return False, "Unusually long"
    
    common_placeholders = ['N/A', 'None', 'Unknown', 'TBD', 'To Be Determined']
    if old_concept.lower() in (placeholder.lower() for placeholder in common_placeholders):
        return False, f"Common placeholder: {old_concept}"
    
    if old_concept.lower() == name.lower():
        return False, "Identical to Names entry"
    
    return True, "Valid"

def process_chunk_removals(chunk: pd.DataFrame) -> Tuple[Dict[str, Set[str]], List[Tuple[str, str, str, str]]]:
    grouped = defaultdict(set)
    flagged_entries = []
    
    for _, row in chunk.iterrows():
        code_system, code, name, old_concept, new_concept = preprocess_row_removals(row)
        
        if code_system not in EXCLUDED_CODE_SYSTEMS:
            is_valid, reason = validate_old_concept_removals(old_concept, name)
            if is_valid and old_concept != new_concept:
                grouped[old_concept].add(name)
                logging.debug(f"Added code {code} to concept {old_concept} because it was moved to {new_concept}")
            elif not is_valid:
                flagged_entries.append((old_concept, code_system, code, reason))
                logging.debug(f"Flagged code {code} for concept {old_concept}: {reason}")
            else:
                logging.debug(f"Skipped code {code} for concept {old_concept} because it remains in the same concept")
    
    # Remove any concepts that ended up with no codes
    grouped = {k: v for k, v in grouped.items() if v}
    
    return grouped, flagged_entries

def summarize_concept_removals_with_semaphore(*args, **kwargs):
    with semaphore:
        return summarize_concept_removals(*args, **kwargs)

def summarize_concept_removals(concept: str, names: Set[str], generative_ai_inference_client: GenerativeAiInferenceClient, compartment_id: str) -> str:
    if not names:
        return f"{concept}: No codes were removed from this concept."

    names_text = "\n".join(names)

    prompt = f"""Summarize the following information for the concept '{concept}':
Concept: {concept}
Names:
{names_text}

1. Each concept should be summarized only once, with no duplicates or omissions.
2. For each concept, concisely summarize the values in the 'Names' column that are associated with that concept.
3. Present the summary in the following format: Concept: [preamble] for [code summaries].
4. Rotate through these two preambles only: 'Removed codes', 'Codes were removed specifying'.
5. Do not interpret any of the codes. Only provide a summarization of them.
6. Do not include any references to 'Code System' or 'Code' in your summary.

Example of desired output:
Connective Tissue Disorder: Removed codes for unspecified, localized, and other specified connective tissue disorders, with additional codes specifying the body site affected.

Format the summary as:
{concept}: [Your summary here]"""

    chat_request = CohereChatRequest(
        message=prompt,
        max_tokens=600,
        temperature=0.25,
        frequency_penalty=0,
        top_p=0.75,
        top_k=0
    )

    chat_detail = ChatDetails(
        serving_mode=OnDemandServingMode(
            model_id="ocid1.generativeaimodel.oc1.us-chicago-1.amaaaaaask7dceya7ozidbukxwtun4ocm4ngco2jukoaht5mygpgr6gq2lgq"
        ),
        chat_request=chat_request,
        compartment_id=compartment_id
    )

    @retry_on_throttle
    def make_chat_request(chat_detail):
        return generative_ai_inference_client.chat(chat_detail)

    try:
        chat_response = make_chat_request(chat_detail)

        # Parse the response as JSON
        response_json = json.loads(str(chat_response.data))

        # Extract the summary text
        summary_text = response_json.get('chat_response', {}).get('text', '')

        if summary_text:
            return summary_text.strip()
        else:
            logging.warning(f"No 'text' field found in the response for concept '{concept}'. Full response: {response_json}")
            return f"{concept}: Unable to generate summary."

    except json.JSONDecodeError as e:
        logging.error(f"Error parsing JSON response for concept '{concept}': {e}")
        return f"{concept}: Error occurred while parsing the response."
    except oci.exceptions.ServiceError as e:
        if e.status == 429:
            logging.error(f"Rate limit exceeded even after retries for concept '{concept}': {e}")
        else:
            logging.error(f"Service error for concept '{concept}': {e}")
        return f"{concept}: Error occurred while summarizing this concept."
    except Exception as e:
        logging.error(f"Unexpected error summarizing concept '{concept}': {str(e)}")
        return f"{concept}: Unexpected error occurred while summarizing this concept."

def get_total_rows_removals(file_path: str) -> int:
    try:
        xl = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = xl.active  # Selects the first sheet in the workbook
        total_rows = sheet.max_row - 1  # Subtract 1 to exclude header
        return total_rows
    finally:
        xl.close()
        del xl
        gc.collect()

def write_report_removals(output_file: str, summaries: List[str]) -> None:
    with open(output_file, "w", encoding='utf-8') as f:
        # Write introductory sentence
        f.write("The following concepts had codes removed as described:\n\n")
        
        # Write sorted summaries
        sorted_summaries = sorted(summaries, key=lambda x: x.split(':', 1)[0].lower())
        for summary in sorted_summaries:
            f.write(f"{summary}\n\n")

def process_large_dataset_stream_removals(
    file_path: str,
    sheet_name: str,
    output_file: str,
    flagged_file: str,
    generative_ai_inference_client: GenerativeAiInferenceClient,
    compartment_id: str
) -> None:
    sheet_name = get_sheet_name(file_path)
    total_rows = get_total_rows_removals(file_path)
    data_stream = stream_excel_removals(file_path, sheet_name)
    
    all_data = defaultdict(set)
    all_flagged_entries = []
    deprecated_concepts = set()
    processed_rows = 0

    # Process chunks with progress updates
    with tqdm(total=total_rows, desc="Processing Removals", unit="rows") as pbar:
        for chunk in data_stream:
            chunk_data, flagged_entries = process_chunk_removals(chunk)
            for concept, names in chunk_data.items():
                all_data[concept].update(names)
            all_flagged_entries.extend(flagged_entries)

            chunk_size = len(chunk)
            processed_rows += chunk_size
            pbar.update(chunk_size)

            # Update progress tracker if defined
            progress = min(95, int((processed_rows / total_rows) * 100))
            if 'progress_tracker' in globals():
                progress_tracker.update(progress, f"Processing removals: {processed_rows}/{total_rows}")

    sorted_concepts = dict(sorted(all_data.items()))

    # Generate summaries with semaphore-controlled concurrency
    if 'progress_tracker' in globals():
        progress_tracker.update(96, "Generating removal summaries...")
    summaries = []

    concepts_list = list(sorted_concepts.items())

    for i in range(0, len(concepts_list), batch_size):
        batch = concepts_list[i:i + batch_size]
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_concept = {
                executor.submit(
                    summarize_concept_removals_with_semaphore,  # Use the wrapper function
                    concept,
                    names,
                    generative_ai_inference_client,
                    compartment_id
                ): concept
                for concept, names in batch
            }

            for future in as_completed(future_to_concept):
                concept = future_to_concept[future]
                try:
                    summary = future.result()
                    summaries.append(summary)
                except Exception as e:
                    logging.error(f"Error processing concept {concept}: {str(e)}")
                    # Create a basic summary for failed concepts
                    names = sorted_concepts[concept]
                    summaries.append(f"{concept}: Removed {len(names)} codes")

    # Sort summaries alphabetically
    summaries.sort(key=lambda x: x.split(':', 1)[0].lower())

    if 'progress_tracker' in globals():
        progress_tracker.update(98, "Writing removals report...")
    write_report_removals(output_file, summaries)

    with open(flagged_file, "w", encoding='utf-8') as f:
        f.write("Old Concept Name,Code System,Code,Reason\n")
        for entry in sorted(all_flagged_entries):
            f.write(f"{','.join(entry)}\n")

    if 'progress_tracker' in globals():
        progress_tracker.update(100, "Removals processing complete")

#code relocations summary
def code_relocations(file_path: str, removals_output_file: str):
    """
    Function to handle code relocations and append movement information to the removals output.
    """
    def stream_excel(file_path: str, sheet_name: str = None, chunk_size: int = 1000) -> Iterator[pd.DataFrame]:
        try:
            xlsx = openpyxl.load_workbook(file_path, read_only=True)
            if sheet_name is None:
                sheet_name = xlsx.sheetnames[0]  # Use first sheet if none specified
            sheet = xlsx[sheet_name]
            
            rows = sheet.iter_rows(values_only=True)
            header = next(rows)
            
            chunk = []
            for row in rows:
                chunk.append(row)
                if len(chunk) == chunk_size:
                    df = pd.DataFrame(chunk, columns=header)
                    yield df[['Code System', 'Code', 'Names', 'Old Concept Name', 'New Concept Name']]
                    chunk = []
            
            if chunk:
                df = pd.DataFrame(chunk, columns=header)
                yield df[['Code System', 'Code', 'Names', 'Old Concept Name', 'New Concept Name']]
        finally:
            xlsx.close()
            del xlsx
            gc.collect()

    def process_chunk_for_movements(chunk: pd.DataFrame) -> Tuple[Dict[str, Set[str]], Set[str], Set[str]]:
        movements = defaultdict(set)
        all_concepts = set()
        removed_concepts = set()
        code_to_concepts = defaultdict(set)
        code_to_new_concept = {}
        
        for _, row in chunk.iterrows():
            code_system = row['Code System']
            if code_system not in EXCLUDED_CODE_SYSTEMS:  # Apply same exclusions as other processes
                code, name = row['Code'], row['Names']
                old_concept, new_concept = row['Old Concept Name'], row['New Concept Name']
                
                if pd.notna(old_concept) and pd.isna(new_concept):
                    old_concept_clean = clean_text_removals(str(old_concept))
                    code_to_concepts[code].add(old_concept_clean)
                    all_concepts.add(old_concept_clean)
                    removed_concepts.add(old_concept_clean)
                elif pd.isna(old_concept) and pd.notna(new_concept):
                    new_concept_clean = clean_text_removals(str(new_concept))
                    code_to_new_concept[code] = new_concept_clean
        
        for code, old_concepts in code_to_concepts.items():
            if code in code_to_new_concept:
                new_concept = code_to_new_concept[code]
                for old_concept in old_concepts:
                    movements[old_concept].add(new_concept)
        
        return movements, all_concepts, removed_concepts

    def summarize_movements(concept: str, new_concepts: Set[str]) -> str:
        if not new_concepts:
            return "No codes from this concept were moved to new concepts."
        
        new_concepts_list = sorted(new_concepts)
        if len(new_concepts_list) > 10:
            new_concepts_text = ", ".join(new_concepts_list[:10]) + ", and other concepts"
        elif len(new_concepts_list) == 1:
            new_concepts_text = new_concepts_list[0]
        elif len(new_concepts_list) == 2:
            new_concepts_text = f"{new_concepts_list[0]} and {new_concepts_list[1]}"
        else:
            new_concepts_text = ", ".join(new_concepts_list[:-1]) + f", and {new_concepts_list[-1]}"
        
        return f"Codes in this concept were moved to {new_concepts_text}."

    def process_large_dataset_for_movements(file_path: str, output_file: str):
        sheet_name = get_sheet_name(file_path)
        data_stream = stream_excel(file_path, sheet_name)
        
        progress_tracker.update(0, "Processing relocations...")
        
        all_movements = defaultdict(set)
        all_concepts = set()
        removed_concepts = set()
        processed_chunks = 0
        total_chunks = get_total_rows_additions(file_path) // 1000 + 1  # Estimate chunks
        
        for chunk in data_stream:
            chunk_movements, chunk_concepts, chunk_removed = process_chunk_for_movements(chunk)
            for old_concept, new_concepts in chunk_movements.items():
                all_movements[old_concept].update(new_concepts)
            all_concepts.update(chunk_concepts)
            removed_concepts.update(chunk_removed)
            
            processed_chunks += 1
            progress = min(90, int((processed_chunks / total_chunks) * 100))
            progress_tracker.update(progress, f"Processing relocations: chunk {processed_chunks}/{total_chunks}")

        logging.info(f"Total concepts: {len(all_concepts)}")
        logging.info(f"Concepts with removals: {len(removed_concepts)}")
        logging.info(f"Concepts with movements: {len(all_movements)}")

        # Generate summaries for movements
        progress_tracker.update(95, "Generating relocation summaries...")
        movement_summaries = {concept: summarize_movements(concept, all_movements.get(concept, set())) 
                            for concept in removed_concepts}

        # Read and update the removals output file
        progress_tracker.update(98, "Updating removals report with relocations...")
        try:
            with open(output_file, "r", encoding='utf-8') as f:
                lines = f.readlines()
            
            new_lines = []
            processed_concepts = set()
            for line in lines:
                parts = line.split(":", 1)
                if len(parts) == 2:
                    concept = parts[0].strip()
                    if concept in movement_summaries:
                        # Only append movement summary if not already present
                        existing_text = parts[1].strip()
                        movement_text = movement_summaries[concept]
                        if "moved to" not in existing_text:
                            new_line = f"{concept}: {existing_text} {movement_text}\n"
                        else:
                            new_line = line
                        new_lines.append(new_line)
                        processed_concepts.add(concept)
                    else:
                        new_lines.append(line)
                else:
                    new_lines.append(line)
            
            # Append summaries for any concepts not found in the existing file
            unprocessed_concepts = removed_concepts - processed_concepts
            if unprocessed_concepts:
                new_lines.append("\nAdditional concepts with relocated codes:\n")
                for concept in sorted(unprocessed_concepts):
                    new_lines.append(f"{concept}: {movement_summaries[concept]}\n")
            
            # Write the updated content
            with open(output_file, "w", encoding='utf-8') as f:
                f.writelines(new_lines)
            
            progress_tracker.update(100, "Relocation processing complete")
            logging.info(f"Movement summaries inserted into {output_file}")
        
        except Exception as e:
            logging.error(f"Error updating removals file with movements: {str(e)}")
            progress_tracker.update(100, "Error updating relocations")
            raise

    # Main processing call
    sheet_name = get_sheet_name(file_path)
    process_large_dataset_for_movements(file_path, removals_output_file)

@app.route('/concept_summary', methods=['GET', 'POST'])
def concept_summary():
    if request.method == 'POST':
        selected_options = request.form.getlist('options')
        
        if 'file' not in request.files:
            error = "No file part"
            return render_template('concept_summary.html', error=error)
        
        file = request.files['file']
        
        if file.filename == '':
            error = "No selected file"
            return render_template('concept_summary.html', error=error)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            temp_dir = tempfile.mkdtemp()
            file_path = os.path.join(temp_dir, filename)
            file.save(file_path)

            standardized_file_path = standardize_column_names(file_path)

            try:
                output_files = []
                
                # Process Code Additions
                if 'additions' in selected_options:
                    sheet_name = 'Sheet1'  # Replace with actual sheet name or use `get_sheet_name(file_path)` if applicable
                    additions_output = code_additions(
                        file_path=file_path,
                        sheet_name=sheet_name,
                        generative_ai_inference_client=generative_ai_inference_client,
                        compartment_id=compartment_id
                    )
                    output_files.append(additions_output)

                
                # Process Code Removals
                if 'removals' in selected_options:
                    sheet_name = 'Sheet1'  # Replace with actual sheet name or use `get_sheet_name(file_path)` if applicable
                    removals_output = code_removals(
                        file_path=file_path,
                        sheet_name=sheet_name,
                        generative_ai_inference_client=generative_ai_inference_client,
                        compartment_id=compartment_id
                    )
                    output_files.append(removals_output)
                
                # Process Code Relocations
                if 'relocations' in selected_options:
                    # Ensure code_removals has been run
                    if 'removals' not in selected_options:
                        removals_output = code_removals(file_path)
                        output_files.append(removals_output)
                    code_relocations(file_path, removals_output)

                # Define the path for the final combined output file
                combined_output_path = os.path.join(temp_dir, 'combined_output.txt')
                
                # Combine the outputs from the selected options into a single file
                with open(combined_output_path, 'w') as combined_file:
                    for output_file in output_files:
                        with open(output_file, 'r') as f:
                            combined_file.write(f.read())
                            combined_file.write("\n")  # Add a newline between different outputs

                # Read the content of the combined file
                with open(combined_output_path, 'r') as f:
                    content = f.read()

                # Clean up files before sending response
                try:
                    shutil.rmtree(temp_dir)
                except PermissionError:
                    # If cleanup fails, schedule it for later
                    @after_this_request
                    def cleanup(response):
                        try:
                            shutil.rmtree(temp_dir)
                        except (PermissionError, FileNotFoundError):
                            pass  # If cleanup still fails, let the OS handle it
                        return response

                # Create response with the file content
                response = make_response(content)
                response.headers['Content-Type'] = 'text/plain'
                response.headers['Content-Disposition'] = f'attachment; filename=combined_output.txt'
                return response
                
            except Exception as e:
                # If any error occurs during processing, ensure cleanup
                try:
                    shutil.rmtree(temp_dir)
                except (PermissionError, FileNotFoundError):
                    pass  # If cleanup fails, let the OS handle it
                raise e
                
        else:
            error = "Invalid file type. Please upload an Excel file."
            return render_template('concept_summary.html', error=error)
    else:
        return render_template('concept_summary.html')

#concept to concept summary
def extract_concept_name(df, workbook_name):
    # Attempt to extract the high-level concept name from 'Concept Name' column
    if 'Concept Name' in df.columns and not df['Concept Name'].dropna().empty:
        concept_text = df['Concept Name'].dropna().iloc[0]
        concept_name = concept_text.strip()
        if concept_name:
            return concept_name
    # Fallback to workbook name only if no concept name found
    return workbook_name.strip()

def extract_text_for_summary(text: str) -> str:
    """Extract relevant text from Name(s) column for summarization"""
    if not isinstance(text, str):
        return ""

    lines = text.strip().split('\n')

    asterisk_match = re.findall(r'\*(.*?)\*', text)
    if asterisk_match:
        return asterisk_match[0].strip()

    for line in reversed(lines):
        parentheses_match = re.match(r'(.*)\s+\([^)]+\)$', line)
        if parentheses_match:
            return parentheses_match.group(1).strip()

    return lines[-1].strip()

def get_unique_codes(df1, df2, workbook_names):
    """Identify unique and common codes between two dataframes"""
    # Create unique identifiers for each code
    df1['identifier'] = df1['Code System'] + '|' + df1['Code']
    df2['identifier'] = df2['Code System'] + '|' + df2['Code']
    
    # Get sets of identifiers
    set1 = set(df1['identifier'])
    set2 = set(df2['identifier'])
    
    # Find unique and common codes
    unique_to_1 = set1 - set2
    unique_to_2 = set2 - set1
    common = set1 & set2
    
    # Create dataframes for each category
    unique_1_df = df1[df1['identifier'].isin(unique_to_1)][['Code System', 'Code']].copy()
    unique_2_df = df2[df2['identifier'].isin(unique_to_2)][['Code System', 'Code']].copy()
    common_df = df1[df1['identifier'].isin(common)][['Code System', 'Code']].copy()
    
    return unique_1_df, unique_2_df, common_df

def get_concept_summary(
    df: pd.DataFrame,
    workbook_name: str,
    concept_name: str,
    generative_ai_inference_client: GenerativeAiInferenceClient,
    compartment_id: str
) -> dict:
    """
    Prepare summary data for a concept using Oracle Generative AI Inference Client.
    """
    clean_workbook_name = workbook_name.strip()
    concept_name = concept_name.strip()
    
    # Extract active names using extract_text_for_summary on each 'Name(s)' entry
    active_names = df['Name(s)'].dropna().apply(extract_text_for_summary).unique().tolist()
    active_names = [name for name in active_names if name]
    
    # Combine active names into a single text
    names_text = "\n".join(active_names)
    
    # Check conditions for using the AI model
    use_ai_model = False
    if len(names_text) > 250:
        use_ai_model = True
    else:
        # Check number of concepts with Included = 'Y'
        if 'Included' in df.columns:
            included_concepts = df[df['Included'].str.upper() == 'Y']['Concept Name'].nunique()
            if included_concepts <= 3:
                use_ai_model = True
    
    if use_ai_model:
        try:
            # Prepare prompt
            prompt = f"""Summarize the following information for the concept '{concept_name}':

{names_text}

1. Begin by listing out every unique value in the 'Concept Name' column. Every unique concept should be listed only once, with no duplicates or omissions.
2. For each concept identified in step 1, concisely summarize the values in the 'Name(s)' column that are associated with that concept.
3. Present the summary in the following format: {concept_name}: [Your summary here].
4. Do not interpret any of the codes. Only provide a summarization of them.
5. Do not include any references to 'Code System' or 'Code' in your summary.

Example Input:
Concepts, Codes
Colon Cancer, Stage I Colon Cancer AJCC v8
Colon Cancer, Stage I Colon Cancer AJCC v6 and v7
Colon Cancer, Ca transverse colon
Colon Cancer, Ca descending colon
Colon Cancer, Ca ascending colon
Colon Cancer, Ca splenic flexure - colon
Colon Cancer, Ca sigmoid colon

Desired AI Output:
Colon Cancer: Includes codes related to colon cancer of various regions of the colon, including transverse, ascending, descending, sigmoid colon, and splenic flexure. Additionally, codes related to staging by various systems such as AJCC (American Joint Committee on Cancer).

Format the summary as:
{concept_name}: [Your summary here]"""

            # Set up chat_request
            chat_request = CohereChatRequest(
                message=prompt,
                max_tokens=600,
                temperature=0.25,
                frequency_penalty=0,
                top_p=0.75,
                top_k=0
            )
            
            # Set up chat_detail
            chat_detail = ChatDetails(
                serving_mode=OnDemandServingMode(
                    model_id="ocid1.generativeaimodel.oc1.us-chicago-1.amaaaaaask7dceya7ozidbukxwtun4ocm4ngco2jukoaht5mygpgr6gq2lgq"
                ),
                chat_request=chat_request,
                compartment_id=compartment_id
            )
            
            # Call the Generative AI Inference Client
            chat_response = generative_ai_inference_client.chat(chat_detail)
            
            # Parse the JSON response
            response_json = json.loads(str(chat_response.data))
            
            # Extract the text from the parsed JSON
            summary_text = response_json.get('chat_response', {}).get('text', '')
            
            if summary_text:
                code_summaries = summary_text.strip()
            else:
                print(f"No 'text' field found in the response.")
                code_summaries = f"{concept_name}: Unable to generate summary."
        except Exception as e:
            print(f"Generative AI Inference API error: {str(e)}")
            code_summaries = f"{concept_name}: Error occurred while summarizing this concept."
    else:
        # Fallback if AI model is not used
        if not names_text.startswith(f"{concept_name}:"):
            code_summaries = f"{concept_name}: Includes codes for {', '.join(active_names)}"
        else:
            code_summaries = names_text
    
    return {
        'workbook_name': clean_workbook_name,
        'concept_name': concept_name,
        'code_summaries': code_summaries,
        'full_title': f"{clean_workbook_name} {concept_name}"
    }

def write_excel_section(writer, title, df, start_row):
    """Helper function to write a section to Excel with proper formatting"""
    # Use title verbatim
    section_title = f"Unique to {title}" if title != "Codes in Both Concepts" else "Codes in Both Concepts"
    
    # Write section title
    pd.DataFrame({
        'A': [section_title],
    }).to_excel(writer, sheet_name='Code Differences', startrow=start_row, index=False, header=False)
    
    # Write column headers
    pd.DataFrame({
        'A': ['Code System'],
        'B': ['Code']
    }).to_excel(writer, sheet_name='Code Differences', startrow=start_row+1, index=False, header=False)
    
    # Write data if exists
    if not df.empty:
        df.to_excel(writer, sheet_name='Code Differences', startrow=start_row+2, index=False, header=False)
        end_row = start_row + len(df) + 2
    else:
        end_row = start_row + 2
    
    # Add empty row after each section
    pd.DataFrame({'A': [''], 'B': ['']}).to_excel(writer, sheet_name='Code Differences', startrow=end_row+1, index=False, header=False)
    
    return end_row + 2  # Update start_row for the next section

def chunk_text(text: str, max_length: int = 1800) -> list:
    """Split text into chunks that won't exceed Cohere's limits"""
    words = text.split(', ')
    chunks = []
    current_chunk = []
    current_length = 0
    
    for word in words:
        if current_length + len(word) + 2 > max_length:  # +2 for ', '
            if current_chunk:
                chunks.append(', '.join(current_chunk))
            current_chunk = [word]
            current_length = len(word)
        else:
            current_chunk.append(word)
            current_length += len(word) + 2
    
    if current_chunk:
        chunks.append(', '.join(current_chunk))
    
    return chunks

@app.route('/concept_compare', methods=['GET', 'POST'])
def concept_compare():
    if request.method == 'POST':
        if 'files' not in request.files:
            return render_template('concept_compare.html', error="No files uploaded")
        
        files = request.files.getlist('files')
        
        if not files or any(file.filename == '' for file in files):
            return render_template('concept_compare.html', error="No files selected")

        if not all(allowed_file(file.filename) for file in files):
            return render_template('concept_compare.html', error="Invalid file type. Please upload Excel files only.")

        try:
            temp_dir = tempfile.mkdtemp()
            dfs = []
            workbook_names = []

            # Process each uploaded file
            for file in files:
                filename = secure_filename(file.filename)
                filepath = os.path.join(temp_dir, filename)
                file.save(filepath)
                
                df = pd.read_excel(filepath)
                workbook_name = os.path.splitext(filename)[0]
                workbook_names.append(workbook_name)
                
                if 'Included' in df.columns:
                    df = df[df['Included'].str.upper() == 'Y']
                
                required_cols = ['Code System', 'Code', 'Name(s)', 'Concept Name']
                df = df[required_cols]
                
                # Handle multiple codes per concept
                df['Code System'] = df['Code System'].fillna('').astype(str)
                df['Code'] = df['Code'].fillna('').astype(str)
                
                # Split and explode code systems and codes
                for col in ['Code System', 'Code']:
                    for sep in [', ', ',', '\n', ';']:
                        if df[col].str.contains(sep).any():
                            df[col] = df[col].str.split(sep)
                            break
                    else:
                        df[col] = df[col].apply(lambda x: [x])
                
                df = df.explode('Code System').explode('Code')
                
                # Clean up the data
                df['Code System'] = df['Code System'].str.strip()
                df['Code'] = df['Code'].str.strip()
                
                # Store original Names for summary
                df[f'{workbook_name} Name(s)'] = df['Name(s)']
                
                dfs.append(df)

            # Create separate dataframes for each workbook
            df1, df2 = dfs[0], dfs[1]

            # Extract concept names
            concept_name1 = extract_concept_name(df1, workbook_names[0])
            concept_name2 = extract_concept_name(df2, workbook_names[1])

            # Get unique and common codes
            unique_1_df, unique_2_df, common_df = get_unique_codes(df1, df2, workbook_names)

            # Get summaries with concept names
            summary1 = get_concept_summary(
                df1, workbook_names[0], concept_name1,
                generative_ai_inference_client, compartment_id
            )
            summary2 = get_concept_summary(
                df2, workbook_names[1], concept_name2,
                generative_ai_inference_client, compartment_id
            )

            # Save to Excel with multiple sheets
            output_path = os.path.join(temp_dir, 'concept_comparison.xlsx')
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Initialize start_row
                start_row = 0

                # Write unique codes for first workbook, using workbook name verbatim
                start_row = write_excel_section(writer, workbook_names[0], unique_1_df, start_row)

                # Write unique codes for second workbook, using workbook name verbatim
                start_row = write_excel_section(writer, workbook_names[1], unique_2_df, start_row)

                # Write common codes
                start_row = write_excel_section(writer, 'Codes in Both Concepts', common_df, start_row)

                # Prepare data for the 'Summaries' sheet
                summary_data = []

                for summary in [summary1, summary2]:
                    # Add the title (workbook name + concept name)
                    summary_data.append([summary['full_title']])
                    # Add the summary text on the next row
                    summary_data.append([summary['code_summaries']])
                    summary_data.append([''])  # Empty row between summaries

                # Write summaries to 'Summaries' sheet
                summary_df = pd.DataFrame(summary_data, columns=['Summaries'])
                summary_df.to_excel(writer, sheet_name='Summaries', startrow=0, index=False, header=False)

                # Left-justify the text in 'Summaries' sheet
                worksheet = writer.sheets['Summaries']
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='left')

                # Auto-adjust column widths
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column_cells in worksheet.columns:
                        length = max(len(str(cell.value)) for cell in column_cells)
                        worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

            try:
                response = send_file(
                    output_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='concept_comparison.xlsx'
                )
                return response
            except Exception as e:
                print(f"Error sending file: {str(e)}")
                return render_template('concept_compare.html', error=f"Error sending file: {str(e)}")

        except Exception as e:
            print(f"Error: {str(e)}")
            return render_template('concept_compare.html', error=f"An error occurred: {str(e)}")
        
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # Return template for GET request
    return render_template('concept_compare.html')

#compare two datasets
@app.route('/compare_datasets', methods=['GET', 'POST'])
def compare_datasets():
    if request.args.get('download_template'):
        # Create template file
        wb = openpyxl.Workbook()
        
        # Set up Sheet1
        sheet1 = wb.active
        sheet1.title = 'Sheet1'
        sheet1['A1'] = 'Dataset 1 Values'
        
        # Format A column as text
        for cell in sheet1['A']:
            cell.number_format = '@'
        
        # Create and set up Sheet2
        sheet2 = wb.create_sheet('Sheet2')
        sheet2['A1'] = 'Dataset 2 Values'
        
        # Format A column as text
        for cell in sheet2['A']:
            cell.number_format = '@'
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            template_path = temp_file.name
            wb.save(template_path)
            
            return send_file(
                template_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='dataset_comparison_template.xlsx'
            )

    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('compare_datasets.html', error='No file uploaded')
        
        file = request.files['file']
        if file.filename == '':
            return render_template('compare_datasets.html', error='No file selected')
        
        if not file.filename.endswith('.xlsx'):
            return render_template('compare_datasets.html', error='Please upload an Excel (.xlsx) file')
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        output_path = temp_file.name
        temp_file.close()

        try:
            # Save uploaded file to temporary location
            file.save(output_path)
            
            # Process the file
            wb = openpyxl.load_workbook(output_path)
            
            # Verify required sheets exist
            if 'Sheet1' not in wb.sheetnames or 'Sheet2' not in wb.sheetnames:
                os.unlink(output_path)
                return render_template('compare_datasets.html', 
                    error='File must contain Sheet1 and Sheet2')
            
            # Get the sheets
            sheet1 = wb['Sheet1']
            sheet2 = wb['Sheet2']
            
            # Create Sheet3 if it doesn't exist
            if 'Sheet3' not in wb.sheetnames:
                sheet3 = wb.create_sheet('Sheet3')
            else:
                sheet3 = wb['Sheet3']
            
            # Format column A as text in both sheets
            for cell in sheet1['A']:
                cell.number_format = '@'
            for cell in sheet2['A']:
                cell.number_format = '@'
            
            # Get codes from Sheet1 and Sheet2, ensuring all are strings
            codes1 = set(str(cell.value) for cell in sheet1['A'] if cell.value is not None)
            codes2 = set(str(cell.value) for cell in sheet2['A'] if cell.value is not None)
            
            # Find unique codes
            unique_to_sheet1 = codes1 - codes2
            unique_to_sheet2 = codes2 - codes1
            
            # Clear Sheet3
            for row in sheet3['A1:B'+str(sheet3.max_row)]:
                for cell in row:
                    cell.value = None
            
            # Write results to Sheet3
            sheet3['A1'] = "Values in Sheet1 not in Sheet2"
            sheet3['B1'] = "Values in Sheet2 not in Sheet1"
            
            # Write unique values and format as text
            for i, code in enumerate(sorted(unique_to_sheet1), start=2):
                cell = sheet3[f'A{i}']
                cell.value = code
                cell.number_format = '@'
                
            for i, code in enumerate(sorted(unique_to_sheet2), start=2):
                cell = sheet3[f'B{i}']
                cell.value = code
                cell.number_format = '@'
            
            # Save and close the workbook
            wb.save(output_path)
            wb.close()
            
            # Return the processed file
            return_data = send_file(
                output_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'compared_{secure_filename(file.filename)}'
            )
            
            # Add cleanup callback
            @return_data.call_on_close
            def cleanup():
                try:
                    if os.path.exists(output_path):
                        os.unlink(output_path)
                except Exception:
                    pass  # Ignore cleanup errors
                    
            return return_data
            
        except Exception as e:
            # Clean up in case of error
            try:
                if os.path.exists(output_path):
                    os.unlink(output_path)
            except Exception:
                pass
            return render_template('compare_datasets.html', 
                error=f'An error occurred while processing the file: {str(e)}')
    
    return render_template('compare_datasets.html')

#personalized code compare
@app.route('/personalized_code_compare', methods=['GET', 'POST'])
def personalized_code_compare():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('personalized_code_compare.html', 
                                 error='No file uploaded')
        
        file = request.files['file']
        if file.filename == '':
            return render_template('personalized_code_compare.html', 
                                 error='No file selected')
        
        if not file.filename.endswith('.xlsx'):
            return render_template('personalized_code_compare.html', 
                                 error='Please upload an Excel (.xlsx) file')
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        output_path = temp_file.name
        temp_file.close()

        try:
            # Save uploaded file
            file.save(output_path)
            
            # Read the sheets
            try:
                concepts_df = pd.read_excel(output_path, sheet_name='Sheet1')
                sheet2_df = pd.read_excel(output_path, sheet_name='Sheet2')
            except Exception as e:
                return render_template('personalized_code_compare.html', 
                    error='Error reading sheets. Please ensure Sheet1 and Sheet2 exist and are properly formatted.')
            
            # Validate Sheet1 has Concept column
            if 'Concept' not in concepts_df.columns:
                return render_template('personalized_code_compare.html', 
                    error='Sheet1 must have a "Concept" column')

            # Process the data
            # Ensure concept names are strings
            concepts_df['Concept'] = concepts_df['Concept'].astype(str)
            concepts_to_find = concepts_df['Concept'].dropna().tolist()

            # Ensure sheet2 columns are strings
            sheet2_df['Old Concept Name'] = sheet2_df['Old Concept Name'].astype(str)

            # Initialize results
            results_df = pd.DataFrame()
            match_status = {}
            old_to_new_concepts = {}

            # Process each concept
            for concept in concepts_to_find:
                # Find matching rows
                matching_rows = sheet2_df[sheet2_df['Old Concept Name'].str.contains(concept, na=False, case=False)]
                
                # Update match status
                match_status[concept] = 'matches identified' if not matching_rows.empty else 'matches not identified'
                
                if not matching_rows.empty:
                    # Process matching rows
                    for _, row in matching_rows.iterrows():
                        code_system = row['Code System']
                        code = row['Code']
                        old_concept_name = row['Old Concept Name']
                        old_concept_alias = row['Old Concept Aliases']
                        
                        key = (old_concept_name, old_concept_alias)
                        if key not in old_to_new_concepts:
                            old_to_new_concepts[key] = set()
                        
                        # Find related rows
                        related_rows = sheet2_df[
                            (sheet2_df['Code System'] == code_system) &
                            (sheet2_df['Code'] == code)
                        ]
                        
                        # Update results
                        results_df = pd.concat([results_df, related_rows], ignore_index=True)
                        
                        # Collect new concepts
                        for _, related_row in related_rows.iterrows():
                            new_name = related_row['New Concept Name']
                            new_alias = related_row['New Concept Aliases']
                            new_name = new_name if pd.notna(new_name) else ''
                            new_alias = new_alias if pd.notna(new_alias) else ''
                            old_to_new_concepts[key].add((new_name, new_alias))

            # Remove duplicates
            results_df.drop_duplicates(inplace=True)
            
            # Update concept match status
            concepts_df['Match Status'] = concepts_df['Concept'].map(match_status)
            
            # Create Sheet4 data
            sheet4_rows = [
                {
                    'Old Concept Name': old_name,
                    'Old Concept Alias': old_alias,
                    'New Concept Name': new_name,
                    'New Concept Alias': new_alias
                }
                for (old_name, old_alias), new_concepts in old_to_new_concepts.items()
                for new_name, new_alias in sorted(new_concepts)
            ]
            sheet4_df = pd.DataFrame(sheet4_rows)
            
            # Write results to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                concepts_df.to_excel(writer, sheet_name='Sheet1', index=False)
                results_df.to_excel(writer, sheet_name='Sheet3', index=False)
                sheet4_df.to_excel(writer, sheet_name='Sheet4', index=False)
            
            # Return the processed file
            return_data = send_file(
                output_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'processed_{secure_filename(file.filename)}'
            )
            
            # Add cleanup callback
            @return_data.call_on_close
            def cleanup():
                try:
                    if os.path.exists(output_path):
                        os.unlink(output_path)
                except Exception:
                    pass
                    
            return return_data
            
        except Exception as e:
            # Clean up in case of error
            try:
                if os.path.exists(output_path):
                    os.unlink(output_path)
            except Exception:
                pass
            return render_template('personalized_code_compare.html', 
                error=f'An error occurred while processing the file: {str(e)}')
    
    return render_template('personalized_code_compare.html')

#----embedding----
# Load OCI configuration
try:
    config_embed = oci.config.from_file(OCI_CONFIG_PATH, OCI_CONFIG_PROFILE)
    logging.info("OCI configuration loaded successfully.")
except Exception as e:
    logging.error(f"Error loading OCI configuration: {str(e)}")
    exit(1)

# Initialize GenAI inference client
generative_ai_inference_client_embed = GenerativeAiInferenceClient(
    config=config_embed,
    service_endpoint=SERVICE_ENDPOINT,
    retry_strategy=oci.retry.NoneRetryStrategy(),
    timeout=(10, 240)
)

# Use tenancy OCID as the compartment ID
compartment_id_embed = config_embed['tenancy']

# Function to check allowed file
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_for_summary_embed(text: str) -> str:
    """Extract relevant text from Name(s) column for summarization"""
    if not isinstance(text, str):
        return ""
    lines = text.strip().split('\n')
    # First check for text between asterisks
    asterisk_match = re.findall(r'\*(.*?)\*', text)
    if asterisk_match:
        return asterisk_match[0].strip()
    # Then check for specific patterns
    for line in reversed(lines):
        # Look for text before parentheses - (a la "finding")
        parentheses_match = re.match(r'(.*?)\s*\([^)]+\)*$', line)
        if parentheses_match:
            return parentheses_match.group(1).strip()
    return lines[-1].strip()

def identify_condition_qualifiers_embed(text):
    """Identify detailed medical qualifiers and modifiers"""
    qualifiers_dict = {
        'severity': ['mild', 'moderate', 'severe', 'critical'],
        'timing': ['acute', 'chronic', 'early', 'late'],
        'specificity': ['primary', 'secondary'],
        'complications': ['with', 'without', 'complicated', 'uncomplicated'],
        'symptoms': ['proteinuria', 'edema', 'seizures'],
        'specific_conditions': ['hellp', 'eclampsia', 'pre-eclampsia', 'gestational'],
        'characteristics': ['essential', 'malignant', 'benign']
    }

    found_qualifiers = {}
    text_lower = text.lower()

    for category, quals in qualifiers_dict.items():
        for qualifier in quals:
            if re.search(rf'\b{qualifier}\b', text_lower):
                if category not in found_qualifiers:
                    found_qualifiers[category] = []
                found_qualifiers[category].append(qualifier)

    return found_qualifiers

def create_combined_text_embed(row):
    """Create rich context combining condition name and qualifiers"""
    qualifiers_flat = []
    for category, quals in row['qualifiers'].items():
        qualifiers_flat.extend(quals)

    # Prioritize specific conditions in the combined text
    specific_conditions = row['qualifiers'].get('specific_conditions', [])
    if specific_conditions:
        combined_text = f"{' '.join(specific_conditions)} {row['extracted_name']}"
    else:
        combined_text = row['extracted_name']

    # Add other qualifiers
    other_qualifiers = [q for q in qualifiers_flat if q not in specific_conditions]
    if other_qualifiers:
        combined_text += f" {' '.join(other_qualifiers)}"

    return combined_text

def perform_clustering_embed(embeddings, min_clusters=8, max_clusters=15):
    """Perform clustering with optimal number of clusters"""
    embeddings = np.nan_to_num(embeddings.astype(float))

    best_silhouette = -1
    best_n_clusters = min_clusters
    best_labels = None

    for n_clusters in range(min_clusters, max_clusters + 1):
        kmeans = KMeans(n_clusters=n_clusters, random_state=42)
        try:
            labels = kmeans.fit_predict(embeddings)
            if len(set(labels)) > 1:
                silhouette_avg = silhouette_score(embeddings, labels)
                if silhouette_avg > best_silhouette:
                    best_silhouette = silhouette_avg
                    best_n_clusters = n_clusters
                    best_labels = labels
        except Exception as e:
            logging.error(f"Error in clustering with {n_clusters} clusters: {e}")
            continue

    return best_labels if best_labels is not None else np.zeros(len(embeddings), dtype=int)

def load_and_embed_concepts_embed(concept_list, model):
    """
    Create embeddings for existing concept names to use as potential cluster labels
    """
    logging.info("Generating embeddings for existing concepts...")
    concept_embeddings = model.encode(concept_list, show_progress_bar=True)
    return concept_embeddings, concept_list

async def generate_cluster_names_with_concepts_embed(df, cluster_ids, existing_concepts, model, batch_size=5):
    """
    Enhanced cluster name generation that considers existing concepts as style references
    """
    # Calculate mean embeddings for each cluster
    logging.info("Calculating mean embeddings for each cluster...")
    cluster_mean_embeddings = []
    for cluster_id in cluster_ids:
        cluster_mask = df['cluster'] == cluster_id
        cluster_texts = df[cluster_mask]['combined_text'].tolist()
        cluster_embeddings = model.encode(cluster_texts)
        mean_embedding = np.mean(cluster_embeddings, axis=0)
        cluster_mean_embeddings.append(mean_embedding)

    # Get concept embeddings
    concept_embeddings, concept_list = load_and_embed_concepts_embed(existing_concepts, model)

    cluster_names = {}
    total_batches = (len(cluster_ids) + batch_size - 1) // batch_size

    # Select a sample of existing concepts to use as style examples
    concept_samples = np.random.choice(existing_concepts, size=min(10, len(existing_concepts)), replace=False)
    concept_samples_text = ' | '.join(concept_samples)

    with tqdm(total=total_batches, desc="Generating cluster names", unit="batch") as pbar:
        for i in range(0, len(cluster_ids), batch_size):
            batch = cluster_ids[i:i + batch_size]
            tasks = [
                process_cluster_with_concepts_embed(
                    df,
                    cluster_id,
                    concept_samples_text  # Pass the concept samples as style reference
                )
                for cluster_id in batch
            ]
            batch_results = await asyncio.gather(*tasks)
            for cluster_id, name in batch_results:
                cluster_names[cluster_id] = name
            pbar.update(1)

    return cluster_names

async def process_cluster_with_concepts_embed(df, cluster_id, concept_samples_text):
    """
    Process a single cluster, using concept samples as style reference
    """
    cluster_mask = df['cluster'] == cluster_id
    cluster_df = df[cluster_mask]

    if len(cluster_df) == 0:
        return cluster_id, f"Cluster {cluster_id}"

    cluster_texts = cluster_df['extracted_name'].tolist()

    # Aggregate qualifiers for the cluster
    cluster_qualifiers = {}
    for quals in cluster_df['qualifiers']:
        for category, values in quals.items():
            if category not in cluster_qualifiers:
                cluster_qualifiers[category] = []
            cluster_qualifiers[category].extend(values)

    # Count occurrences and keep most common qualifiers
    for category in cluster_qualifiers:
        qualifier_counts = Counter(cluster_qualifiers[category])
        cluster_qualifiers[category] = [
            qual for qual, count in qualifier_counts.most_common(3)
            if count >= len(cluster_texts) * 0.2
        ]

    # Modified prompt to include concept samples as style reference
    prompt = f"""
You are a medical condition classifier focusing on specific subtypes and manifestations.

**Example Conditions in this Cluster:** {' | '.join(cluster_texts[:5])}
**Common Qualifiers Found:** {'; '.join([f"{cat}: {', '.join(quals)}"
                                     for cat, quals in cluster_qualifiers.items()
                                     if quals])}

**Existing Concept Names (for reference):** {concept_samples_text}

Task: Create a specific medical cluster name that:
- Accurately represents the conditions in the cluster.
- Is similar in style and terminology to the existing concept names provided.
- Uses precise medical terminology.

Cluster Name (2-5 words):
"""

    chat_request = CohereChatRequest(
        message=prompt,
        max_tokens=15,
        temperature=0.2,
        frequency_penalty=0,
        top_p=0.75,
        top_k=0
    )

    chat_detail = ChatDetails(
        serving_mode=OnDemandServingMode(
            model_id=GENERATIVE_AI_MODEL_ID
        ),
        chat_request=chat_request,
        compartment_id=compartment_id_embed
    )

    try:
        response = await asyncio.to_thread(generative_ai_inference_client_embed.chat, chat_detail)
        response_json = json.loads(str(response.data))
        generated_text = response_json.get('chat_response', {}).get('text', '').strip()
        if not generated_text:
            generated_text = f"Cluster {cluster_id}"
        return cluster_id, generated_text
    except Exception as e:
        logging.error(f"Error generating cluster name: {e}")
        return cluster_id, f"Cluster {cluster_id}"

def create_concept_template_embed(output_path='concept_dictionary.xlsx'):
    """
    Create a template Excel file for managing medical concepts
    """
    template_df = pd.DataFrame({
        'concept_name': ['Hypertensive Disorder', 'Chronic Kidney Disease'],  # Example concepts
    })

    try:
        template_df.to_excel(output_path, index=False)
        logging.info(f"Created concept template at {output_path}")

        # Upload the template to OCI
        object_storage_client = oci.object_storage.ObjectStorageClient(config_embed)
        namespace = NAMESPACE
        bucket_name = BUCKET_NAME
        with open(output_path, 'rb') as file_data:
            object_storage_client.put_object(
                namespace,
                bucket_name,
                'concept_dictionary.xlsx',
                file_data
            )
        logging.info(f"Uploaded concept template to OCI Object Storage")
        return True
    except Exception as e:
        logging.error(f"Error creating or uploading concept template: {e}")
        return False

def load_concepts_embed():
    """
    Load concepts from OCI Object Storage

    Returns:
        list: List of concept names
    """
    try:
        # Initialize Object Storage client
        logging.debug("Initializing Object Storage client.")
        object_storage_client = oci.object_storage.ObjectStorageClient(config_embed)
        namespace = NAMESPACE
        logging.debug(f"Using namespace: {namespace}")

        bucket_name = BUCKET_NAME
        object_name = 'concept_dictionary.xlsx'
        logging.debug(f"Attempting to get object '{object_name}' from bucket '{bucket_name}'.")

        # Get the object
        response = object_storage_client.get_object(namespace, bucket_name, object_name)
        logging.info(f"Successfully retrieved '{object_name}' from bucket '{bucket_name}'.")

        # Read the content directly from response.data
        content = response.data.content
        
        # Create BytesIO object from the content
        bytes_io = io.BytesIO(content)

        # Read Excel
        df = pd.read_excel(bytes_io)
        logging.debug(f"Excel columns: {df.columns.tolist()}")

        required_columns = ['concept_name']

        # Verify required columns exist
        if not all(col in df.columns for col in required_columns):
            logging.error(f"Missing required columns. File must contain: {required_columns}")
            return []

        concepts = df['concept_name'].dropna().tolist()
        logging.info(f"Loaded {len(concepts)} concepts from '{object_name}'.")
        return concepts

    except oci.exceptions.ServiceError as e:
        logging.error(f"ServiceError while loading concepts: {e}")
        if e.status == 404:
            logging.warning(f"Concept file '{object_name}' not found in bucket '{bucket_name}'. Creating template...")
            # Create a template and upload to OCI
            if create_concept_template_embed():
                logging.info("Template created and uploaded successfully.")
            else:
                logging.error("Failed to create and upload template.")
            return []
        else:
            return []
    except Exception as e:
        logging.error(f"Unexpected error while loading concepts: {e}", exc_info=True)
        return []
    
def save_concept_report_embed(df, output_path):
    """
    Save the clustered report to OCI Object Storage

    Args:
        df (pd.DataFrame): DataFrame containing clustered data
        output_path (str): Local path to save the Excel file temporarily
    """
    try:
        df_sorted = df.sort_values('cluster_name')
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_df = pd.DataFrame({
                'Cluster Name': df_sorted['cluster_name'].unique(),
                'Count': df_sorted.groupby('cluster_name').size(),
                'Sample Conditions': df_sorted.groupby('cluster_name')['extracted_name']
                    .apply(lambda x: ' | '.join(x.iloc[:3])),
                'Common Qualifiers': df_sorted.groupby('cluster_name')['qualifiers']
                    .apply(lambda x: ' | '.join(set(str(qual) for quals in x for qual in quals.values()))),
            }).reset_index(drop=True)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            for cluster_name, group in df_sorted.groupby('cluster_name'):
                # Ensure the sheet name is within Excel's limit and valid
                sheet_name = re.sub(r'[^\w\s-]', '', cluster_name)[:31]
                group.to_excel(writer, sheet_name=sheet_name, index=False)

        # Upload to OCI Object Storage
        object_storage_client = oci.object_storage.ObjectStorageClient(config_embed)
        
        # Create a simpler filename without duplicating "clustered_codes_report"
        object_name = f'clustered_codes_report_{Path(output_path).stem.replace("clustered_codes_report_", "")}.xlsx'
        
        with open(output_path, 'rb') as file_data:
            object_storage_client.put_object(
                namespace_name=NAMESPACE,  # Changed from namespace to namespace_name
                bucket_name=BUCKET_NAME,
                object_name=object_name,
                put_object_body=file_data
            )

        logging.info(f"Uploaded report to OCI Object Storage as {object_name}")
        return object_name

    except Exception as e:
        logging.error(f"Error saving report: {e}")
        return None
    
async def process_file_embed(file_path, concepts):
    # Load and filter data
    logging.info("Loading and filtering data...")
    df = pd.read_excel(file_path)
    df = df[df['Included'] == 'Y']
    df = df[~df['Code System'].isin({'Read Codes v2', 'Read Codes v3', 'ICD-10-SE'})]

    # Load existing concepts
    existing_concepts = concepts
    if not existing_concepts:
        logging.warning("No concepts loaded. Please populate concept_dictionary.xlsx")
        return None, None

    logging.info(f"Loaded {len(existing_concepts)} concepts for matching")

    # Process text and generate combined text for richer embedding context
    logging.info("Processing text data...")
    df['extracted_name'] = df['Name(s)'].apply(extract_text_for_summary_embed)
    df['qualifiers'] = df['extracted_name'].apply(identify_condition_qualifiers_embed)
    df['combined_text'] = df.apply(create_combined_text_embed, axis=1)

    # Generate embeddings using SentenceTransformer
    logging.info("Generating embeddings...")
    model = SentenceTransformer('all-MiniLM-L6-v2')
    embeddings = model.encode(df['combined_text'].tolist(), show_progress_bar=True)
    embeddings = np.array(embeddings)

    # Dimensionality reduction
    logging.info("Performing dimensionality reduction...")
    pca = PCA(n_components=50)
    reduced_embeddings = pca.fit_transform(embeddings)

    # Perform clustering
    logging.info("Performing clustering...")
    df['cluster'] = perform_clustering_embed(reduced_embeddings)

    # Generate cluster names using existing concepts as style reference
    logging.info("Generating cluster names...")
    cluster_names = await generate_cluster_names_with_concepts_embed(
        df,
        df['cluster'].unique(),
        existing_concepts,
        model
    )
    df['cluster_name'] = df['cluster'].map(cluster_names)

    # Visualization
    logging.info("Creating visualization...")
    tsne = TSNE(n_components=2, random_state=42)
    df[['x', 'y']] = tsne.fit_transform(reduced_embeddings)

    fig = px.scatter(
        df, x='x', y='y', color='cluster_name',
        hover_data=['extracted_name', 'Code'],
        title='Cluster Visualization'
    )
    plot_div = pio.to_html(fig, full_html=False)

    # Export results
    logging.info("Exporting results...")
    report_filename = f"clustered_codes_report_{Path(file_path).stem}.xlsx"
    local_report_path = os.path.join(app.config['RESULTS_FOLDER'], report_filename)
    oci_report_object = save_concept_report_embed(df, local_report_path)

    return plot_div, oci_report_object

@app.route('/concept_embedding', methods=['GET', 'POST'])
def concept_embedding_embed():
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # If user does not select file, browser may submit an empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            uploaded_file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(uploaded_file_path)
            logging.info(f"File saved to {uploaded_file_path}")

            # Load concepts from OCI
            concepts_embed = load_concepts_embed()
            if not concepts_embed:
                flash('No concepts loaded. Please ensure concept_dictionary.xlsx exists in OCI Object Storage.')
                return redirect(request.url)

            # Process the file asynchronously
            plot_div, report_object = asyncio.run(process_file_embed(uploaded_file_path, concepts_embed))

            if plot_div and report_object:
                # Generate URLs to access the report
                plot_filename = f"plot_{Path(uploaded_file_path).stem}.html"
                plot_path = os.path.join(app.config['RESULTS_FOLDER'], plot_filename)
                
                # Write the plot with UTF-8 encoding
                with open(plot_path, 'w', encoding='utf-8') as f:
                    f.write(plot_div)

                # Remove the uploaded file after processing
                os.remove(uploaded_file_path)

                return render_template('concept_embedding_embed.html',
                                    results_ready=True,
                                    plot_file=plot_filename,
                                    excel_file=report_object)
            else:
                flash('Error processing the file.')
                return redirect(request.url)
        else:
            flash('Allowed file types are .xlsx')
            return redirect(request.url)
    else:
        return render_template('concept_embedding_embed.html', results_ready=False)

@app.route('/results/<filename>')
def uploaded_file_embed(filename):
    return send_from_directory(app.config['RESULTS_FOLDER'], filename)

@app.route('/static/<path:filename>')
def static_files_embed(filename):
   return send_from_directory('static', filename)

@app.route('/download_report/<filename>')
def download_report_embed(filename):
    """
    Route to download the Excel report from OCI Object Storage
    """
    try:
        # Initialize Object Storage client
        object_storage_client = oci.object_storage.ObjectStorageClient(config_embed)
        
        # Get the object from OCI
        response = object_storage_client.get_object(
            namespace_name=NAMESPACE,  # Changed from namespace to namespace_name
            bucket_name=BUCKET_NAME,
            object_name=filename
        )
        
        # Create a response with the file content
        return Response(
            response.data.content,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except oci.exceptions.ServiceError as e:
        logging.error(f"OCI Service Error downloading report: {e.message}")
        abort(404)
    except Exception as e:
        logging.error(f"Error downloading report: {str(e)}")
        abort(404)

#----pivot table to codes in concept----
@app.route('/find_missing_codes', methods=['GET', 'POST'])
def find_missing_codes():
    if request.method == 'POST':
        try:
            if 'file' not in request.files:
                return render_template('find_missing_codes.html', error='No file uploaded')
            
            file = request.files['file']
            if file.filename == '':
                return render_template('find_missing_codes.html', error='No file selected')

            if not file.filename.endswith('.xlsx'):
                return render_template('find_missing_codes.html', error='Please upload an Excel (.xlsx) file')

            # Read the Excel file
            df_sheet1 = pd.read_excel(file, sheet_name='Sheet1')
            df_sheet2 = pd.read_excel(file, sheet_name='Sheet2')

            # Initialize lists for codes and descriptions
            codes = []
            descriptions = []
            current_code = None
            current_descriptions = []
            
            # Regular expression pattern for codes (only digits)
            code_pattern = re.compile(r'^\d+$')

            # Process Sheet1
            for value in df_sheet1['Codes and Descriptions']:
                value = str(value).strip()
                if code_pattern.match(value):
                    if current_code is not None:
                        codes.append(current_code)
                        descriptions.append('; '.join(current_descriptions))
                    current_code = value
                    current_descriptions = []
                else:
                    current_descriptions.append(value)

            # Add the last code and descriptions
            if current_code is not None:
                codes.append(current_code)
                descriptions.append('; '.join(current_descriptions))

            # Create processed DataFrame
            df_sheet1_processed = pd.DataFrame({
                'Code': codes,
                'Description': descriptions
            })

            # Ensure code columns are string type
            df_sheet1_processed['Code'] = df_sheet1_processed['Code'].astype(str).str.strip()
            df_sheet2['Code'] = df_sheet2['Code'].astype(str).str.strip()

            # Find missing codes
            mask = ~df_sheet1_processed['Code'].isin(df_sheet2['Code'])
            df_sheet3 = df_sheet1_processed.loc[mask, ['Code', 'Description']]

            # Create output file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_sheet1_processed.to_excel(writer, sheet_name='Sheet1', index=False)
                df_sheet2.to_excel(writer, sheet_name='Sheet2', index=False)
                df_sheet3.to_excel(writer, sheet_name='Sheet3', index=False)
            
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='missing_codes_report.xlsx'
            )

        except Exception as e:
            return render_template('find_missing_codes.html', error=str(e))

    return render_template('find_missing_codes.html')

#placeholders
@app.route('/measure_summary', methods=['GET', 'POST'])
def measure_summary():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('measure_summary.html', error="No file uploaded")
        
        # TODO: Add summary logic
        return render_template('measure_summary.html', error="Feature coming soon!")
    
    return render_template('measure_summary.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)